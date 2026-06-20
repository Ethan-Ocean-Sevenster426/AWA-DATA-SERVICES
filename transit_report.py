#!/usr/bin/env python3
"""
AWA Data Services - Transit External Customer Report
====================================================
Pulls the "External Customer Report MOC" from CargoWise TWD (WiseGrid) for the
configured branches, builds an Excel workbook matching the master template, and
(optionally) uploads it to SharePoint via Microsoft Graph.

Runs unattended on Linux (systemd timer / cron / Docker / GitHub Actions).
All configuration comes from environment variables - NO secrets are stored in code.

Filters replicated from the saved search "External Customer Report MOC":
  * RCN reference     : has NO word starting with "s00"
  * Unloaded time     : a package unloaded in the last N months (REPORT_MONTHS)
  * Package status    : NONE of the packages are DEP (departed)
  * Booking party     : has NO word starting with "ISCM"
"""
import os, sys, json, re, logging, calendar, datetime, collections
import urllib.request, urllib.parse, urllib.error, http.cookiejar

# --------------------------------------------------------------------------- #
# Config (environment-driven; tiny .env loader so local dev works without deps)
# --------------------------------------------------------------------------- #
def _load_dotenv(path=".env"):
    if os.path.exists(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())

_load_dotenv()

def _env(key, default=None, required=False):
    val = os.environ.get(key, default)
    if required and not val:
        sys.exit(f"FATAL: required environment variable {key} is not set")
    return val

# CargoWise / WiseGrid
CW_BASE        = _env("CW_BASE_URL", "https://www-isbint.wisegrid.net").rstrip("/")
CW_MODULE      = _env("CW_MODULE", "TWD")
ODATA_MODEL    = _env("CW_ODATA_MODEL", "TransitWarehouse")
CW_USERNAME    = _env("CW_USERNAME", required=True)
CW_PASSWORD    = _env("CW_PASSWORD", required=True)
BRANCH_CODES   = [b.strip().upper() for b in _env("CW_BRANCH_CODES", "DOR,CON").split(",") if b.strip()]
DEPARTMENT_CODE= _env("CW_DEPARTMENT_CODE", "BRN").upper()
REPORT_MONTHS  = int(_env("REPORT_MONTHS", "12"))

# Output / behaviour
OUTPUT_DIR     = _env("OUTPUT_DIR", "./output")
OUTPUT_FILE    = _env("SHAREPOINT_FILENAME", "RCN_External_Customer_Report_MOC.xlsx")
DO_UPLOAD      = _env("UPLOAD", "true").lower() in ("1", "true", "yes")

# Microsoft Graph / SharePoint
AZ_TENANT      = _env("AZURE_TENANT_ID")
AZ_CLIENT      = _env("AZURE_CLIENT_ID")
AZ_SECRET      = _env("AZURE_CLIENT_SECRET")
SP_HOST        = _env("SHAREPOINT_HOSTNAME")
SP_SITE        = _env("SHAREPOINT_SITE_PATH")           # e.g. /sites/DataPrime
SP_FOLDER      = _env("SHAREPOINT_FOLDER", "")          # e.g. External Report Transit Pull

log = logging.getLogger("transit_report")

# --------------------------------------------------------------------------- #
# HTTP helper (cookie-aware) + small utilities
# --------------------------------------------------------------------------- #
_cookies = http.cookiejar.CookieJar()
_opener  = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_cookies))

def _http(url, data=None, headers=None, method=None, timeout=180):
    req = urllib.request.Request(url, data=data, headers=headers or {}, method=method)
    return _opener.open(req, timeout=timeout)

def months_ago_iso(months):
    t = datetime.date.today()
    idx = t.month - 1 - months
    year = t.year + idx // 12
    month = idx % 12 + 1
    day = min(t.day, calendar.monthrange(year, month)[1])
    return f"{year:04d}-{month:02d}-{day:02d}T00:00:00Z"

def has_word_starting(text, prefix):
    if not text:
        return False
    p = prefix.lower()
    return any(w.startswith(p) for w in re.split(r"[^A-Za-z0-9]+", text.lower()) if w)

def parse_dt(s):
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})", s)  # wall-clock component
    if not m:
        return None
    try:
        return datetime.datetime(*[int(x) for x in m.groups()])
    except ValueError:
        return None

# --------------------------------------------------------------------------- #
# CargoWise client
# --------------------------------------------------------------------------- #
class CargoWise:
    def __init__(self):
        self.auth = f"{CW_BASE}/Glow/auth/v2"
        self.odata = f"{CW_BASE}/Glow/odata/{ODATA_MODEL}"
        self.user_key = None
        self.branch_map = {}      # CODE -> branchKey
        self.dept_key = None

    def _post(self, path, body):
        data = json.dumps(body).encode()
        r = _http(f"{self.auth}/{path}", data=data, method="POST",
                  headers={"Content-Type": "application/json", "Accept": "application/json",
                           "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop", "Origin": CW_BASE})
        return json.load(r)

    def _get(self, resource, params):
        url = f"{self.odata}/{resource}?{urllib.parse.urlencode(params)}"
        r = _http(url, headers={"Accept": "application/json", "wtg-app": "Glow",
                                "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop"})
        return json.load(r)

    def login(self):
        res = self._post("credential/claim/Staff",
                         {"userName": CW_USERNAME, "password": CW_PASSWORD, "setTokenCookie": True})
        if res.get("result") != 0:
            sys.exit(f"FATAL: CargoWise login failed (result={res.get('result')})")
        self.user_key = res["userKey"]
        log.info("Authenticated to CargoWise as %s", res.get("userDisplayName"))
        self._resolve_contexts()

    def _resolve_contexts(self):
        res = self._post("credential/context/list",
                         {"logonProviderType": "Staff", "userKey": self.user_key, "useTokenCookie": True})
        for b in res.get("branchInfos", []):
            if b["code"].upper() in BRANCH_CODES:
                self.branch_map[b["code"].upper()] = b["key"]
        for d in res.get("departmentInfos", []):
            if d["code"].upper() == DEPARTMENT_CODE:
                self.dept_key = d["key"]
        missing = [c for c in BRANCH_CODES if c not in self.branch_map]
        if missing:
            sys.exit(f"FATAL: branch code(s) not found for this user: {missing}")
        if not self.dept_key:
            sys.exit(f"FATAL: department code {DEPARTMENT_CODE} not found for this user")
        # establish an initial branch session so reference lookups (OData) work
        self._select_branch(self.branch_map[BRANCH_CODES[0]])

    def _select_branch(self, branch_key):
        self._post("credential/context/select",
                   {"logonProviderType": "Staff", "userKey": self.user_key,
                    "branchKey": branch_key, "departmentKey": self.dept_key, "useAndSetTokenCookie": True})
        self._post("session/begin", {"tokenType": 1, "sessionType": "General", "useAndSetTokenCookie": True})

    def service_levels(self):
        try:
            res = self._get("RefServiceLevelInfos", {"$select": "RS_Code,RS_Description", "$top": "500"})
            return {r["RS_Code"]: r["RS_Description"] for r in res.get("value", [])}
        except Exception as e:                       # non-fatal: fall back to bare codes
            log.warning("Could not load service levels (%s); using codes as-is", e)
            return {}

    def pull_branch(self, branch_key, cutoff_iso):
        self._select_branch(branch_key)
        expand = ("Addresses($expand=Address($expand=OrgHeader,Country)),"
                  "WhsItemPackageStates($select=WPS_Status;$expand=Package($select=KP_Weight,KP_Volume))")
        filt = (f"not startswith(tolower(WRC_ConsignmentID),'s00') and "
                f"WhsItemPackageStates/any(p: p/WPS_UnloadedTime ge {cutoff_iso}) and "
                f"not WhsItemPackageStates/any(p: p/WPS_Status eq 'DEP')")
        rows, skip, page = [], 0, 50
        while True:
            params = {"$filter": filt, "$expand": expand, "$top": str(page),
                      "$skip": str(skip), "$orderby": "WRC_JobID"}
            for _ in range(4):
                try:
                    d = self._get("WhsItemReceiveConsignments", params)
                    break
                except urllib.error.HTTPError as e:
                    if e.code == 401:                # session expired - re-auth and retry
                        log.info("  session expired, re-authenticating")
                        self.login(); self._select_branch(branch_key); continue
                    raise
            else:
                raise RuntimeError("repeated failures pulling consignments")
            batch = d.get("value", [])
            rows.extend(batch)
            if len(batch) < page:
                break
            skip += page
        return rows

# --------------------------------------------------------------------------- #
# Report shaping (matches master template "External Report.xlsx")
# --------------------------------------------------------------------------- #
# Sheet 1 - "External Report" (matches master template)
EXT_COLUMNS = ["Branch", "Receive Consignment ID", "Closed", "RCN Reference", "Consignor",
               "Consignee", "Booking Party", "Expected Arrival at Warehouse",
               "Expected Dispatch from Warehouse", "Next Discharge Port", "Service Level",
               "BKD", "In Warehouse", "DEP", "Overs", "Number of Packages", "Column1"]
# Sheet 2 - "Receive Consignment" (operational view: created/user + weight/volume)
RC_COLUMNS = ["Branch", "Created Time", "Closed", "Receive Consignment ID", "RCN Reference",
              "Number of Packages", "BKD", "In Warehouse", "DEP", "Booking Party",
              "Total Weight", "Total Volume", "Consignor", "Consignee", "Service Level",
              "Create User Code"]
DATE_COLS = {"Closed", "Expected Arrival at Warehouse", "Expected Dispatch from Warehouse", "Created Time"}
WIDTHS = {"Branch": 8, "Receive Consignment ID": 22.6, "Closed": 15.6, "Created Time": 15.6,
          "RCN Reference": 14.9, "Consignor": 40, "Consignee": 40, "Booking Party": 55,
          "Expected Arrival at Warehouse": 28.5, "Expected Dispatch from Warehouse": 32.5,
          "Next Discharge Port": 19, "Service Level": 18, "BKD": 8, "In Warehouse": 12,
          "DEP": 8, "Overs": 8, "Number of Packages": 18, "Column1": 10,
          "Total Weight": 13, "Total Volume": 13, "Create User Code": 14}
INWHS = {"ARV", "PUT", "PIC", "CTT", "STA", "FLO", "REC", "RCV"}

def _addr(rec, atype):
    for a in rec.get("Addresses", []):
        if a.get("E2_AddressType") == atype:
            ad = a.get("Address") or {}
            oh = ad.get("OrgHeader") or {}
            ctry = ad.get("Country") or {}
            name = oh.get("OH_FullName") or ad.get("OA_CompanyNameOverride") or ""
            a1 = ad.get("OA_Address1") or ad.get("OA_Code") or ""
            a2 = ad.get("OA_Address2") or ""
            csz = " ".join(x for x in (ad.get("OA_City") or "", ad.get("OA_State") or "",
                                       ad.get("OA_PostCode") or "") if x)
            country = ctry.get("RN_Desc") or ad.get("OA_RN_NKCountryCode") or ""
            full = ", ".join(x for x in (name, a1, a2, csz, country) if x)
            return full, oh.get("OH_Code") or "", name
    return "", "", ""

def _counts(rec):
    ps = rec.get("WhsItemPackageStates", [])
    c = collections.Counter(p.get("WPS_Status") for p in ps)
    weight = sum((p.get("Package") or {}).get("KP_Weight") or 0 for p in ps)
    volume = sum((p.get("Package") or {}).get("KP_Volume") or 0 for p in ps)
    return (c.get("BKD", 0), sum(v for k, v in c.items() if k in INWHS), c.get("DEP", 0),
            sum(c.values()), round(weight, 3), round(volume, 3))

def shape_rows(branch, records, svc_map):
    out = []
    for r in records:
        bp_full, bp_code, bp_name = _addr(r, "BKD")
        if has_word_starting(bp_name, "ISCM") or has_word_starting(bp_code, "ISCM"):
            continue
        if has_word_starting(r.get("WRC_ConsignmentID") or "", "s00"):
            continue
        cnr, _, _ = _addr(r, "CRG")
        cne, _, _ = _addr(r, "CED")
        bkd, inw, dep, tot, weight, volume = _counts(r)
        sl = r.get("WRC_RS_NKServiceLevel") or ""
        sl = f"{sl} - {svc_map[sl]}" if sl in svc_map else sl
        out.append({
            "Branch": branch,
            "Receive Consignment ID": r.get("WRC_JobID"),
            "Created Time": parse_dt(r.get("WRC_SystemCreateTimeUtc")),
            "Closed": parse_dt(r.get("WRC_CompleteTime")),
            "RCN Reference": r.get("WRC_ConsignmentID"),
            "Consignor": cnr, "Consignee": cne, "Booking Party": bp_full,
            "Expected Arrival at Warehouse": parse_dt(r.get("WRC_ExpectedArrivalTime")),
            "Expected Dispatch from Warehouse": parse_dt(r.get("WRC_ExpectedDispatchTime")),
            "Next Discharge Port": r.get("WRC_RL_NKNextDischargePort"),
            "Service Level": sl,
            "BKD": bkd, "In Warehouse": inw, "DEP": dep, "Overs": 0,
            "Number of Packages": tot, "Column1": None,
            "Total Weight": weight, "Total Volume": volume,
            "Create User Code": r.get("WRC_SystemCreateUser"),
        })
    return out

def build_workbook(rows, path):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    def add_sheet(wb, title, columns, table_name):
        ws = wb.create_sheet(title)
        ws.append(columns)
        for r in rows:
            ws.append([r.get(c) for c in columns])
        for ci, col in enumerate(columns, 1):
            L = get_column_letter(ci)
            ws.column_dimensions[L].width = WIDTHS.get(col, 12)
            if col in DATE_COLS:
                for cell in ws[L][1:]:
                    cell.number_format = "dd-mmm-yy hh:mm"
        ws.freeze_panes = "A2"
        ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
        tbl = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "External Report", EXT_COLUMNS, "ExternalReport")
    add_sheet(wb, "Receive Consignment", RC_COLUMNS, "ReceiveConsignment")
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)

# --------------------------------------------------------------------------- #
# SharePoint upload (Microsoft Graph, app-only client-credentials)
# --------------------------------------------------------------------------- #
def upload_to_sharepoint(local_path):
    for k, v in {"AZURE_TENANT_ID": AZ_TENANT, "AZURE_CLIENT_ID": AZ_CLIENT,
                 "AZURE_CLIENT_SECRET": AZ_SECRET, "SHAREPOINT_HOSTNAME": SP_HOST,
                 "SHAREPOINT_SITE_PATH": SP_SITE}.items():
        if not v:
            sys.exit(f"FATAL: UPLOAD=true but {k} is not set")
    graph = "https://graph.microsoft.com/v1.0"

    body = urllib.parse.urlencode({"client_id": AZ_CLIENT, "client_secret": AZ_SECRET,
                                   "scope": "https://graph.microsoft.com/.default",
                                   "grant_type": "client_credentials"}).encode()
    tok = json.load(_http(f"https://login.microsoftonline.com/{AZ_TENANT}/oauth2/v2.0/token",
                          data=body, method="POST",
                          headers={"Content-Type": "application/x-www-form-urlencoded"}))["access_token"]
    H = {"Authorization": "Bearer " + tok}

    site = json.load(_http(f"{graph}/sites/{SP_HOST}:{SP_SITE}", headers=H))
    drive = json.load(_http(f"{graph}/sites/{site['id']}/drive", headers=H))
    dest = "/".join(p for p in (SP_FOLDER, os.path.basename(local_path)) if p)
    url = f"{graph}/drives/{drive['id']}/root:/{urllib.parse.quote(dest)}:/content"
    with open(local_path, "rb") as f:
        content = f.read()
    res = json.load(_http(url, data=content, method="PUT",
                          headers={**H, "Content-Type": "application/octet-stream"}))
    log.info("Uploaded to SharePoint: %s (%s bytes)", res.get("webUrl"), res.get("size"))

# --------------------------------------------------------------------------- #
def main():
    logging.basicConfig(level=logging.INFO, stream=sys.stdout,
                        format="%(asctime)s %(levelname)s %(message)s")
    cutoff = months_ago_iso(REPORT_MONTHS)
    log.info("Transit External Report - branches=%s, unloaded since %s", BRANCH_CODES, cutoff)

    cw = CargoWise()
    cw.login()
    svc_map = cw.service_levels()

    all_rows = []
    for code in BRANCH_CODES:
        records = cw.pull_branch(cw.branch_map[code], cutoff)
        rows = shape_rows(code, records, svc_map)
        log.info("  %s: %d consignments", code, len(rows))
        all_rows.extend(rows)

    out_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE)
    build_workbook(all_rows, out_path)
    log.info("Workbook written: %s (%d rows)", out_path, len(all_rows))

    if DO_UPLOAD:
        upload_to_sharepoint(out_path)
    else:
        log.info("UPLOAD disabled; skipping SharePoint upload")
    log.info("Done.")

if __name__ == "__main__":
    main()
