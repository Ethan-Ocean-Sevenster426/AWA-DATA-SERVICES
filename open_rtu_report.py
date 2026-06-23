#!/usr/bin/env python3
"""
AWA Data Services - Open RTU's Report
=====================================
Pulls the "OPEN RTU'S" saved search (Receive Transportation Units that have been
created but not yet unloaded) from CargoWise TWD (WiseGrid) for both branches
(DOR + CON), and writes a single "Receive Transportation Unit" sheet matching the
master "Open RTU's.xlsx", then (optionally) uploads to SharePoint via Graph.

Saved search "OPEN RTU'S" (module IEntityInfo_IWhsItemReceiveTransportationUnit):
  * CREATETIME          HasDate     -> WRH_SystemCreateTimeUtc ne null
  * UNLOADCOMPLETETIME  HasNoDate   -> WRH_UnloadCompleteTime  eq null
Both conditions are reproduced exactly (validated: returns the same 6 CON rows the
live grid shows).

Timezone: "Created Time" (actual) is converted true-instant -> display tz (UTC-6),
matching the master file (verified TR00055946 16:09Z -> 10:09). "Unload Complete"
is blank by definition for open RTUs.

Config is environment-driven (see .env.example). No secrets are stored in code.
"""
import os, sys, json, re, logging, datetime
import urllib.request, urllib.parse, urllib.error, http.cookiejar

def _load_dotenv(path=".env"):
    if os.path.exists(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())
_load_dotenv()

def _env(key, default=None, required=False):
    val = os.environ.get(key, default)
    if required and not val:
        sys.exit(f"FATAL: required environment variable {key} is not set")
    return val

CW_BASE      = _env("CW_BASE_URL", "https://www-isbint.wisegrid.net").rstrip("/")
CW_MODULE    = _env("CW_MODULE", "TWD")
ODATA_MODEL  = _env("CW_ODATA_MODEL", "TransitWarehouse")
CW_USERNAME  = _env("CW_USERNAME", required=True)
CW_PASSWORD  = _env("CW_PASSWORD", required=True)
BRANCH_CODES = [b.strip().upper() for b in _env("OPENRTU_BRANCH_CODES", "DOR,CON").split(",") if b.strip()]
DEPT_CODE    = _env("CW_DEPARTMENT_CODE", "BRN").upper()
DISPLAY_TZ_OFFSET = float(_env("CC_DISPLAY_TZ_OFFSET", "-6"))   # CargoWise fixed display offset

OUTDIR       = _env("OUTPUT_DIR", "./output")
OUT_FILE     = _env("OPENRTU_FILENAME", "Open RTU's.xlsx")
DO_UPLOAD    = _env("UPLOAD", "true").lower() in ("1", "true", "yes")

AZ_TENANT  = _env("AZURE_TENANT_ID"); AZ_CLIENT = _env("AZURE_CLIENT_ID"); AZ_SECRET = _env("AZURE_CLIENT_SECRET")
SP_HOST    = _env("SHAREPOINT_HOSTNAME"); SP_SITE = _env("SHAREPOINT_SITE_PATH")
SP_FOLDER  = _env("OPENRTU_SHAREPOINT_FOLDER", _env("RTU_SHAREPOINT_FOLDER", "RTU Report"))

COLUMNS = ["Branch", "Receive Transportation Unit ID", "RTU Reference", "Transport Company",
           "Load Type", "Staging Location", "Drivers Name", "Expected Number of Packages",
           "Unloaded Packages", "Number of Handling Units", "Unload Complete", "Created Time"]
DATE_COLS = {"Unload Complete", "Created Time"}

log = logging.getLogger("open_rtu_report")
_cookies = http.cookiejar.CookieJar()
_opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_cookies))

def _http(url, data=None, headers=None, method=None, timeout=180):
    return _opener.open(urllib.request.Request(url, data=data, headers=headers or {}, method=method), timeout=timeout)

def parse_dt(s):
    """Parse CargoWise DateTimeOffset, normalise to the fixed display offset, truncate to minute."""
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})(?::(\d{2}))?(?:\.\d+)?(Z|[+-]\d{2}:?\d{2})?", s)
    if not m:
        return None
    y, mo, d, h, mi = (int(x) for x in m.groups()[:5])
    se = int(m.group(6) or 0); off = m.group(7)
    try:
        dt = datetime.datetime(y, mo, d, h, mi, se)
    except ValueError:
        return None
    if off and off != "Z":
        sign = 1 if off[0] == "+" else -1
        dt -= datetime.timedelta(hours=sign * int(off[1:3]), minutes=sign * int(off[-2:]))
    dt += datetime.timedelta(hours=DISPLAY_TZ_OFFSET)
    return dt.replace(second=0, microsecond=0)

class CargoWise:
    def __init__(self):
        self.auth = f"{CW_BASE}/Glow/auth/v2"; self.odata = f"{CW_BASE}/Glow/odata/{ODATA_MODEL}"
        self.user_key = None; self.branch_map = {}; self.dept_key = None

    def _post(self, path, body):
        r = _http(f"{self.auth}/{path}", data=json.dumps(body).encode(), method="POST",
                  headers={"Content-Type": "application/json", "Accept": "application/json",
                           "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop", "Origin": CW_BASE})
        return json.load(r)

    def _get(self, resource, params):
        url = f"{self.odata}/{resource}?{urllib.parse.urlencode(params)}"
        return json.load(_http(url, headers={"Accept": "application/json", "wtg-app": "Glow",
                                             "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop"}))

    def login(self):
        res = self._post("credential/claim/Staff",
                         {"userName": CW_USERNAME, "password": CW_PASSWORD, "setTokenCookie": True})
        if res.get("result") != 0:
            sys.exit(f"FATAL: CargoWise login failed (result={res.get('result')})")
        self.user_key = res["userKey"]
        ctx = self._post("credential/context/list",
                         {"logonProviderType": "Staff", "userKey": self.user_key, "useTokenCookie": True})
        for b in ctx.get("branchInfos", []):
            if b["code"].upper() in BRANCH_CODES:
                self.branch_map[b["code"].upper()] = b["key"]
        for d in ctx.get("departmentInfos", []):
            if d["code"].upper() == DEPT_CODE:
                self.dept_key = d["key"]
        missing = [c for c in BRANCH_CODES if c not in self.branch_map]
        if missing:
            sys.exit(f"FATAL: branch code(s) not found: {missing}")
        if not self.dept_key:
            sys.exit(f"FATAL: department {DEPT_CODE} not found")
        self._select(self.branch_map[BRANCH_CODES[0]])
        log.info("Authenticated as %s", res.get("userDisplayName"))

    def _select(self, branch_key):
        self._post("credential/context/select",
                   {"logonProviderType": "Staff", "userKey": self.user_key,
                    "branchKey": branch_key, "departmentKey": self.dept_key, "useAndSetTokenCookie": True})
        self._post("session/begin", {"tokenType": 1, "sessionType": "General", "useAndSetTokenCookie": True})

    def country_map(self):
        m, skip = {}, 0
        try:
            while True:
                d = self._get("RefCountryInfos", {"$select": "RN_Code,RN_Desc", "$top": "50", "$skip": str(skip)})
                v = d.get("value", [])
                m.update({r["RN_Code"]: r["RN_Desc"] for r in v})
                if len(v) < 50:
                    break
                skip += 50
        except Exception as e:
            log.warning("Could not load country map (%s)", e)
        return m

    def pull_open_rtus(self, branch_key):
        self._select(branch_key)
        filt = "WRH_SystemCreateTimeUtc ne null and WRH_UnloadCompleteTime eq null"
        expand = ("Addresses($expand=Address($expand=OrgHeader)),"
                  "StagingLocation($select=WLV_LocationString),"
                  "WhsItemPackageStates($count=true;$top=2000;$select=WPS_IsHandlingUnit)")
        sel = ("WRH_ReferenceNumber,WRH_VehicleReference,WRH_UnitType,WRH_SignedBy,"
               "WRH_UnloadCompleteTime,WRH_SystemCreateTimeUtc")
        rows, skip, page = [], 0, 50
        while True:
            params = {"$filter": filt, "$expand": expand, "$select": sel,
                      "$top": str(page), "$skip": str(skip), "$orderby": "WRH_ReferenceNumber"}
            for _ in range(4):
                try:
                    d = self._get("WhsItemReceiveTransportationUnits", params); break
                except urllib.error.HTTPError as e:
                    if e.code == 401:
                        log.info("  session expired, re-authenticating"); self.login(); self._select(branch_key); continue
                    raise
            else:
                raise RuntimeError("repeated failures pulling open RTUs")
            batch = d.get("value", [])
            rows.extend(batch)
            if len(batch) < page:
                break
            skip += page
        return rows

def transport_company(rtu, cmap):
    for a in rtu.get("Addresses", []):
        if a.get("E2_AddressType") == "TRA":
            ad = a.get("Address") or {}; oh = ad.get("OrgHeader") or {}
            name = a.get("E2_CompanyName") or oh.get("OH_FullName") or ad.get("OA_CompanyNameOverride") or ""
            a1 = a.get("E2_Address1") or ad.get("OA_Address1") or ad.get("OA_Code") or ""
            a2 = a.get("E2_Address2") or ad.get("OA_Address2") or ""
            city = a.get("E2_City") or ad.get("OA_City") or ""
            state = a.get("E2_State") or ad.get("OA_State") or ""
            pc = a.get("E2_Postcode") or ad.get("OA_PostCode") or ""
            cc = a.get("E2_RN_NKCountryCode") or ad.get("OA_RN_NKCountryCode") or ""
            country = cmap.get(cc, cc)
            csz = " ".join(x for x in (city, state, pc) if x)
            return ", ".join(x for x in (name, a1, a2, csz, country) if x)
    return ""

def shape(branch, records, cmap):
    out = []
    for r in records:
        ps = r.get("WhsItemPackageStates", [])
        stg = (r.get("StagingLocation") or {}).get("WLV_LocationString") or ""
        out.append({
            "Branch": branch,
            "Receive Transportation Unit ID": r.get("WRH_ReferenceNumber"),
            "RTU Reference": r.get("WRH_VehicleReference"),
            "Transport Company": transport_company(r, cmap),
            "Load Type": r.get("WRH_UnitType"),
            "Staging Location": stg,
            "Drivers Name": r.get("WRH_SignedBy"),
            # Expected = expected packages from a booking/ASN. Open RTUs have none
            # linked (no consignment divots / ASN), so this is 0 - matches the grid.
            "Expected Number of Packages": 0,
            "Unloaded Packages": r.get("WhsItemPackageStates@odata.count", 0),
            "Number of Handling Units": sum(1 for p in ps if p.get("WPS_IsHandlingUnit")),
            "Unload Complete": parse_dt(r.get("WRH_UnloadCompleteTime")),
            "Created Time": parse_dt(r.get("WRH_SystemCreateTimeUtc")),
        })
    return out

def build_workbook(rows, path):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    widths = {"Branch": 8, "Receive Transportation Unit ID": 26, "RTU Reference": 18,
              "Transport Company": 55, "Load Type": 10, "Staging Location": 16, "Drivers Name": 18,
              "Expected Number of Packages": 24, "Unloaded Packages": 16,
              "Number of Handling Units": 22, "Unload Complete": 17, "Created Time": 17}
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Receive Transportation Unit"
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c) for c in COLUMNS])
    for ci, col in enumerate(COLUMNS, 1):
        L = get_column_letter(ci); ws.column_dimensions[L].width = widths.get(col, 12)
        if col in DATE_COLS:
            for cell in ws[L][1:]:
                cell.number_format = "dd-mmm-yy hh:mm"
    ws.freeze_panes = "A2"
    ws.add_table(Table(displayName="OpenRTU",
                       ref=f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}",
                       tableStyleInfo=TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)))
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True); wb.save(path)

def upload(local_path):
    for k, v in {"AZURE_TENANT_ID": AZ_TENANT, "AZURE_CLIENT_ID": AZ_CLIENT, "AZURE_CLIENT_SECRET": AZ_SECRET,
                 "SHAREPOINT_HOSTNAME": SP_HOST, "SHAREPOINT_SITE_PATH": SP_SITE}.items():
        if not v:
            sys.exit(f"FATAL: UPLOAD=true but {k} is not set")
    graph = "https://graph.microsoft.com/v1.0"
    body = urllib.parse.urlencode({"client_id": AZ_CLIENT, "client_secret": AZ_SECRET,
                                   "scope": "https://graph.microsoft.com/.default",
                                   "grant_type": "client_credentials"}).encode()
    tok = json.load(_http(f"https://login.microsoftonline.com/{AZ_TENANT}/oauth2/v2.0/token", data=body,
                          method="POST", headers={"Content-Type": "application/x-www-form-urlencoded"}))["access_token"]
    H = {"Authorization": "Bearer " + tok}
    site = json.load(_http(f"{graph}/sites/{SP_HOST}:{SP_SITE}", headers=H))
    drive = json.load(_http(f"{graph}/sites/{site['id']}/drive", headers=H))
    dest = "/".join(p for p in (SP_FOLDER, os.path.basename(local_path)) if p)
    url = f"{graph}/drives/{drive['id']}/root:/{urllib.parse.quote(dest)}:/content"
    res = json.load(_http(url, data=open(local_path, "rb").read(), method="PUT",
                          headers={**H, "Content-Type": "application/octet-stream"}))
    log.info("Uploaded: %s (%s bytes)", res.get("webUrl"), res.get("size"))

def main():
    logging.basicConfig(level=logging.INFO, stream=sys.stdout, format="%(asctime)s %(levelname)s %(message)s")
    log.info("Open RTU's Report - branches=%s, created has date + unload complete is blank", BRANCH_CODES)
    cw = CargoWise(); cw.login()
    cmap = cw.country_map()
    all_rows = []
    for code in BRANCH_CODES:
        recs = cw.pull_open_rtus(cw.branch_map[code])
        rows = shape(code, recs, cmap)
        log.info("  %s: %d open RTUs", code, len(rows))
        all_rows.extend(rows)
    out_path = os.path.join(OUTDIR, OUT_FILE)
    build_workbook(all_rows, out_path)
    log.info("Workbook written: %s (%d rows)", out_path, len(all_rows))
    if DO_UPLOAD:
        upload(out_path)
    else:
        log.info("UPLOAD disabled; skipping SharePoint upload")
    log.info("Done.")

if __name__ == "__main__":
    main()
