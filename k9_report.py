#!/usr/bin/env python3
"""
AWA Data Services - K9 Line Item Inventory Report (CON)
======================================================
Reproduces the CargoWise "Packages - K9 CON Inspections MOC" saved search
(Transit Warehouse -> Packages) for the CON branch: every package physically in
the warehouse (PUT etc.), EXCLUDING booked/departed/loaded/packing statuses and
the door/build staging locations.

Unlike the other reports, this one MAINTAINS A SINGLE GROWING FILE: each run adds
today's snapshot (one row per package, tagged with a "Inventory Date") to one flat
sheet and re-uploads it. That single table is the Power BI source - Power BI then
builds the date matrix (Shipment Count = distinct RCN, Total Weight KG, Total CBM)
itself. No sheet-per-day.

Re-running on the same day is idempotent: today's rows are replaced, not doubled.

Config is environment-driven (see .env.example). No secrets are stored in code.
"""
import os, sys, json, re, logging, datetime
import urllib.request, urllib.parse, urllib.error, http.cookiejar

def _load_dotenv(path=".env"):
    if os.path.exists(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())
_load_dotenv()

def _env(key, default=None, required=False):
    val = os.environ.get(key, default)
    if required and not val:
        sys.exit(f"FATAL: required environment variable {key} is not set")
    return val

CW_BASE      = _env("CW_BASE_URL", "https://www-isbint.wisegrid.net").rstrip("/")
CW_MODULE    = _env("CW_MODULE", "TWD")
ODATA_MODEL  = _env("CW_ODATA_MODEL", "TransitWarehouse")
CW_USERNAME  = _env("CW_USERNAME", required=True)
CW_PASSWORD  = _env("CW_PASSWORD", required=True)
BRANCH_CODE  = _env("K9_BRANCH_CODE", "CON").upper()
DEPT_CODE    = _env("CW_DEPARTMENT_CODE", "BRN").upper()
DISPLAY_TZ_OFFSET = float(_env("CC_DISPLAY_TZ_OFFSET", "-6"))   # CargoWise fixed display offset

# --- saved-search filter "K9 CON Inspections MOC" -------------------------------
# Statuses that mean "not in the warehouse / not subject to K9 inspection".
EXCLUDE_STATUS = set(s.strip().upper() for s in _env(
    "K9_EXCLUDE_STATUS", "BKD,ADJ,DEP,UPD,FLO,UPN,PKN,PKD,GIN").split(",") if s.strip())
# Door / build / staging pseudo-locations to drop (exact location-string match).
# NOTE: complete this list from the live saved search if the count doesn't match.
EXCLUDE_LOCATIONS = set(s.strip().upper() for s in _env(
    "K9_EXCLUDE_LOCATIONS",
    "BUILD_OLD-1,BUILD_OLD-2,BUILD_OLD-3,BUILD_OLD-4,BUILD_OLD-5,"
    "CON_DOOR1,CON_DOOR2,CON_DOOR3,DOOR-1,DOOR-2,DOOR-12").split(",") if s.strip())

OUTDIR       = _env("OUTPUT_DIR", "./output")
OUT_FILE     = _env("K9_FILENAME", "K9 Line Item Inventory.xlsx")
SHEET_NAME   = _env("K9_SHEET_NAME", "Inventory")
DO_UPLOAD    = _env("UPLOAD", "true").lower() in ("1", "true", "yes")

AZ_TENANT  = _env("AZURE_TENANT_ID"); AZ_CLIENT = _env("AZURE_CLIENT_ID"); AZ_SECRET = _env("AZURE_CLIENT_SECRET")
SP_HOST    = _env("SHAREPOINT_HOSTNAME"); SP_SITE = _env("SHAREPOINT_SITE_PATH")
SP_FOLDER  = _env("K9_SHAREPOINT_FOLDER", "K9")

# Flat table - one row per package per snapshot day. This IS the Power BI source.
COLUMNS = ["Inventory Date", "Package ID", "RCN Reference", "Location", "Receive Gate out Time",
           "Status", "Consignee", "Consignor", "Booking Party", "Volume", "Weight",
           "Volume M3", "Weight KG"]
DATE_COLS = {"Inventory Date", "Receive Gate out Time"}

STATUS_DESC = {"ARV": "Arrived", "PUT": "Putaway", "PIC": "Picked", "STA": "Staged",
               "CTT": "Counted", "REC": "Received", "RCV": "Received"}

# unit -> cubic metres / kilograms conversion for the numeric (sum-able) columns
M3_PER = {"M3": 1.0, "CBM": 1.0, "CI": 1.6387064e-5, "CF": 0.0283168466, "FT3": 0.0283168466}
KG_PER = {"KG": 1.0, "G": 0.001, "LB": 0.45359237, "LBS": 0.45359237, "T": 1000.0, "MT": 1000.0}

log = logging.getLogger("k9_report")
_cookies = http.cookiejar.CookieJar()
_opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_cookies))

def _http(url, data=None, headers=None, method=None, timeout=180):
    return _opener.open(urllib.request.Request(url, data=data, headers=headers or {}, method=method), timeout=timeout)

def parse_dt(s):
    """CargoWise DateTimeOffset -> fixed display offset (UTC-6), truncated to minute."""
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})(?::(\d{2}))?(?:\.\d+)?(Z|[+-]\d{2}:?\d{2})?", s)
    if not m:
        return None
    y, mo, d, h, mi = (int(x) for x in m.groups()[:5])
    se = int(m.group(6) or 0); off = m.group(7)
    try:
        dt = datetime.datetime(y, mo, d, h, mi, se)
    except ValueError:
        return None
    if off and off != "Z":
        sign = 1 if off[0] == "+" else -1
        dt -= datetime.timedelta(hours=sign * int(off[1:3]), minutes=sign * int(off[-2:]))
    dt += datetime.timedelta(hours=DISPLAY_TZ_OFFSET)
    return dt.replace(second=0, microsecond=0)

def fmt_qty(v, unit):
    s = f"{(v or 0):.3f}".rstrip("0").rstrip(".")
    if s in ("", "-0"):
        s = "0"
    return f"{s} {unit}".strip() if unit else s

def to_m3(v, unit):
    m = round((v or 0) * M3_PER.get((unit or "M3").upper(), 1.0), 6)
    return None if m > 50 else m        # >50 m3 for one package = bad source data

def to_kg(v, unit):
    k = round((v or 0) * KG_PER.get((unit or "KG").upper(), 1.0), 3)
    return None if k > 50000 else k      # >50 t for one package = bad source data

def status_text(code):
    if not code:
        return ""
    return f"{code} - {STATUS_DESC[code]}" if code in STATUS_DESC else code

# --------------------------------------------------------------------------- #
# CargoWise client
# --------------------------------------------------------------------------- #
class CargoWise:
    def __init__(self):
        self.auth = f"{CW_BASE}/Glow/auth/v2"; self.odata = f"{CW_BASE}/Glow/odata/{ODATA_MODEL}"
        self.user_key = None; self.branch_key = None; self.dept_key = None

    def _post(self, path, body):
        r = _http(f"{self.auth}/{path}", data=json.dumps(body).encode(), method="POST",
                  headers={"Content-Type": "application/json", "Accept": "application/json",
                           "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop", "Origin": CW_BASE})
        return json.load(r)

    def _get(self, resource, params):
        url = f"{self.odata}/{resource}?{urllib.parse.urlencode(params)}"
        return json.load(_http(url, headers={"Accept": "application/json", "wtg-app": "Glow",
                                             "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop"}))

    def login(self):
        res = self._post("credential/claim/Staff",
                         {"userName": CW_USERNAME, "password": CW_PASSWORD, "setTokenCookie": True})
        if res.get("result") != 0:
            sys.exit(f"FATAL: CargoWise login failed (result={res.get('result')})")
        self.user_key = res["userKey"]
        ctx = self._post("credential/context/list",
                         {"logonProviderType": "Staff", "userKey": self.user_key, "useTokenCookie": True})
        for b in ctx.get("branchInfos", []):
            if b["code"].upper() == BRANCH_CODE:
                self.branch_key = b["key"]
        for d in ctx.get("departmentInfos", []):
            if d["code"].upper() == DEPT_CODE:
                self.dept_key = d["key"]
        if not self.branch_key:
            sys.exit(f"FATAL: branch {BRANCH_CODE} not found for this user")
        if not self.dept_key:
            sys.exit(f"FATAL: department {DEPT_CODE} not found for this user")
        self._select()
        log.info("Authenticated as %s, branch %s", res.get("userDisplayName"), BRANCH_CODE)

    def _select(self):
        self._post("credential/context/select",
                   {"logonProviderType": "Staff", "userKey": self.user_key,
                    "branchKey": self.branch_key, "departmentKey": self.dept_key, "useAndSetTokenCookie": True})
        self._post("session/begin", {"tokenType": 1, "sessionType": "General", "useAndSetTokenCookie": True})

    def pull_consignments(self):
        # Pull CON consignments that have at least one package NOT in an excluded status
        # (i.e. something still in the warehouse). Per-package status/location filtering
        # then happens client-side in shape().
        ne_chain = " and ".join(f"p/WPS_Status ne '{s}'" for s in sorted(EXCLUDE_STATUS))
        filt = f"WhsItemPackageStates/any(p: {ne_chain})" if ne_chain else None
        expand = ("Addresses($expand=Address($expand=OrgHeader,Country)),"
                  "WhsItemPackageStates("
                  "$select=WPS_PK,WPS_Status;"
                  "$expand=Package($select=KP_Weight,KP_WeightUQ,KP_Volume,KP_VolumeUQ;"
                  "$expand=PackageHeader($select=KPH_PackageID)),"
                  "LastLocation($select=WLV_LocationString),"
                  "TransitReceiveHeader($select=WRH_GateOutTime))")
        sel = "WRC_JobID,WRC_ConsignmentID"
        rows, skip, page = [], 0, 50
        while True:
            params = {"$expand": expand, "$select": sel, "$top": str(page),
                      "$skip": str(skip), "$orderby": "WRC_JobID"}
            if filt:
                params["$filter"] = filt
            for _ in range(4):
                try:
                    d = self._get("WhsItemReceiveConsignments", params); break
                except urllib.error.HTTPError as e:
                    if e.code == 401:
                        log.info("  session expired, re-authenticating"); self.login(); continue
                    raise
            else:
                raise RuntimeError("repeated failures pulling consignments")
            batch = d.get("value", [])
            rows.extend(batch)
            if len(batch) < page:
                break
            skip += page
            if skip % 1000 == 0:
                log.info("  ...%d consignments", len(rows))
        return rows

# --------------------------------------------------------------------------- #
# Shaping: explode consignments -> one row per in-warehouse package
# --------------------------------------------------------------------------- #
def _addr_name(rec, atype):
    for a in rec.get("Addresses", []):
        if a.get("E2_AddressType") == atype:
            ad = a.get("Address") or {}
            oh = ad.get("OrgHeader") or {}
            ctry = ad.get("Country") or {}
            name = oh.get("OH_FullName") or ad.get("OA_CompanyNameOverride") or ""
            a1 = ad.get("OA_Address1") or ""
            a2 = ad.get("OA_Address2") or ""
            csz = " ".join(x for x in (ad.get("OA_City") or "", ad.get("OA_State") or "",
                                       ad.get("OA_PostCode") or "") if x)
            country = ctry.get("RN_Desc") or ad.get("OA_RN_NKCountryCode") or ""
            return ", ".join(x for x in (name, a1, a2, csz, country) if x)
    return ""

def shape(records, snapshot_date):
    out = []
    for r in records:
        rcn = r.get("WRC_ConsignmentID")
        consignee = _addr_name(r, "CED")
        consignor = _addr_name(r, "CRG")
        booking   = _addr_name(r, "BKD")
        for p in r.get("WhsItemPackageStates", []):
            st = (p.get("WPS_Status") or "").upper()
            if st in EXCLUDE_STATUS:
                continue
            loc = ((p.get("LastLocation") or {}).get("WLV_LocationString") or "")
            if loc.upper() in EXCLUDE_LOCATIONS:
                continue
            pkg = p.get("Package") or {}
            kph = pkg.get("PackageHeader") or {}
            rtu = p.get("TransitReceiveHeader") or {}
            vol_u = pkg.get("KP_VolumeUQ") or "M3"
            wt_u = pkg.get("KP_WeightUQ") or "KG"
            out.append({
                "Inventory Date": snapshot_date,
                "Package ID": kph.get("KPH_PackageID"),
                "RCN Reference": rcn,
                "Location": loc,
                "Receive Gate out Time": parse_dt(rtu.get("WRH_GateOutTime")),
                "Status": status_text(p.get("WPS_Status")),
                "Consignee": consignee,
                "Consignor": consignor,
                "Booking Party": booking,
                "Volume": fmt_qty(pkg.get("KP_Volume"), vol_u),
                "Weight": fmt_qty(pkg.get("KP_Weight"), wt_u),
                "Volume M3": to_m3(pkg.get("KP_Volume"), vol_u),
                "Weight KG": to_kg(pkg.get("KP_Weight"), wt_u),
            })
    return out

# --------------------------------------------------------------------------- #
# Maintain the single growing workbook
# --------------------------------------------------------------------------- #
def _load_existing(path, snapshot_date):
    """Return prior rows from the master file, dropping any already tagged with
    today's snapshot date (so a re-run replaces today rather than duplicating it)."""
    if not os.path.exists(path):
        return []
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    try:
        header = list(next(it))
    except StopIteration:
        wb.close(); return []
    idx = {h: i for i, h in enumerate(header)}
    di = idx.get("Inventory Date")
    kept = []
    for row in it:
        if not any(c is not None for c in row):
            continue
        rec = {h: row[idx[h]] if idx[h] < len(row) else None for h in COLUMNS if h in idx}
        sd = rec.get("Inventory Date")
        if isinstance(sd, datetime.datetime):
            sd = sd.date()
        if sd == snapshot_date:
            continue                      # drop today's prior rows (idempotent re-run)
        kept.append(rec)
    wb.close()
    return kept

def build_workbook(rows, path):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    wide = {"Inventory Date": 14, "Package ID": 22, "RCN Reference": 14, "Location": 12,
            "Receive Gate out Time": 18, "Status": 16, "Consignee": 45, "Consignor": 45,
            "Booking Party": 45, "Volume": 12, "Weight": 12, "Volume M3": 11, "Weight KG": 11}
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = SHEET_NAME
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c) for c in COLUMNS])
    for ci, col in enumerate(COLUMNS, 1):
        L = get_column_letter(ci); ws.column_dimensions[L].width = wide.get(col, 12)
        if col in DATE_COLS:
            fmt = "dd-mmm-yy" if col == "Inventory Date" else "dd-mmm-yy hh:mm"
            for cell in ws[L][1:]:
                cell.number_format = fmt
    ws.freeze_panes = "A2"
    ws.add_table(Table(displayName="K9Inventory",
                       ref=f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}",
                       tableStyleInfo=TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)))
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True); wb.save(path)

def main():
    logging.basicConfig(level=logging.INFO, stream=sys.stdout, format="%(asctime)s %(levelname)s %(message)s")
    snapshot_date = datetime.date.today()
    log.info("K9 Line Item Inventory - branch=%s, snapshot=%s", BRANCH_CODE, snapshot_date.isoformat())
    cw = CargoWise(); cw.login()
    records = cw.pull_consignments()
    today_rows = shape(records, snapshot_date)
    log.info("  %s: %d packages in K9 inventory today", BRANCH_CODE, len(today_rows))

    out_path = os.path.join(OUTDIR, OUT_FILE)
    # Pull the current master from SharePoint first so we always EXTEND it (history +
    # every prior snapshot), then append/refresh today's rows - never replace.
    if DO_UPLOAD:
        try:
            import sp_upload
            if sp_upload.download(SP_FOLDER, OUT_FILE, out_path):
                log.info("  pulled current master from SharePoint")
            else:
                log.info("  no master on SharePoint yet; starting a fresh one")
        except Exception as e:
            log.warning("  could not pull master (%s); extending local copy instead", e)
    prior = _load_existing(out_path, snapshot_date)
    all_rows = prior + today_rows
    build_workbook(all_rows, out_path)
    log.info("Workbook written: %s (%d rows total, %d days)", out_path, len(all_rows),
             len({r.get("Inventory Date") for r in all_rows}))

    if DO_UPLOAD:
        import sp_upload
        try:
            # single maintained file: no per-run archive copies (SharePoint keeps version history)
            sp_upload.upload(out_path, SP_FOLDER, archive=False)
        except urllib.error.HTTPError as e:
            if e.code == 423:          # file open in Excel/online -> locked; don't fail the run
                log.warning("SharePoint copy is LOCKED (file open in Excel/online?); "
                            "local copy updated, will upload on the next run")
            else:
                raise
    else:
        log.info("UPLOAD disabled; skipping SharePoint upload")
    log.info("Done.")

if __name__ == "__main__":
    main()
