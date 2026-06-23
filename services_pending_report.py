#!/usr/bin/env python3
"""
AWA Data Services - Services Pending (Freight On Hand) Report
============================================================
Pulls the "Services Pending - Freight on-hand" saved search (receive consignments
that have a booked-but-not-completed service AND still have packages in the
warehouse) from CargoWise TWD (WiseGrid) for both branches (DOR + CON), and writes
a single "Receive Consignment" sheet matching the master
"Services Pending - Freight On Hand (Daily).xlsx", then (optionally) uploads to
SharePoint via Microsoft Graph.

Saved search "Services Pending - Freight on-hand" (module IEntityInfo_IWhsItemReceiveConsignment):
  * SERVICEBOOKEDDATE      HasDate     -> a service with ES_BookedDateTimeOffset set
  * SERVICECOMPLETEDDATE   HasNoDate   -> that service has no ES_CompletedDateTimeOffset
  * HASPACKAGESINWAREHOUSE Is true     -> a package in an in-warehouse status
All three reproduced exactly (validated: returns the same 5 CON rows the grid shows).

Package columns (Number of Packages, BKD/ARV/CTT/PIC/PUT, In Warehouse) are the SUM
of package quantities (KP_PackageQty) by status, not the count of records.
Timezone: "Closed" (actual) -> true-instant converted to display tz (UTC-6).

Config is environment-driven (see .env.example). No secrets are stored in code.
"""
import os, sys, json, re, logging, datetime
import urllib.request, urllib.parse, urllib.error, http.cookiejar

def _load_dotenv(path=".env"):
    if os.path.exists(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())
_load_dotenv()

def _env(key, default=None, required=False):
    val = os.environ.get(key, default)
    if required and not val:
        sys.exit(f"FATAL: required environment variable {key} is not set")
    return val

CW_BASE        = _env("CW_BASE_URL", "https://www-isbint.wisegrid.net").rstrip("/")
CW_MODULE      = _env("CW_MODULE", "TWD")
ODATA_MODEL    = _env("CW_ODATA_MODEL", "TransitWarehouse")
CW_USERNAME    = _env("CW_USERNAME", required=True)
CW_PASSWORD    = _env("CW_PASSWORD", required=True)
BRANCH_CODES   = [b.strip().upper() for b in _env("SVP_BRANCH_CODES", "DOR,CON").split(",") if b.strip()]
DEPARTMENT_CODE= _env("CW_DEPARTMENT_CODE", "BRN").upper()
DISPLAY_TZ     = int(_env("SVP_DISPLAY_TZ_OFFSET", _env("CC_DISPLAY_TZ_OFFSET", "-6")))

OUTPUT_DIR     = _env("OUTPUT_DIR", "./output")
OUTPUT_FILE    = _env("SVP_FILENAME", "Services Pending - Freight On Hand (Daily).xlsx")
DO_UPLOAD      = _env("UPLOAD", "true").lower() in ("1", "true", "yes")

AZ_TENANT      = _env("AZURE_TENANT_ID"); AZ_CLIENT = _env("AZURE_CLIENT_ID"); AZ_SECRET = _env("AZURE_CLIENT_SECRET")
SP_HOST        = _env("SHAREPOINT_HOSTNAME"); SP_SITE = _env("SHAREPOINT_SITE_PATH")
SP_FOLDER      = _env("SVP_SHAREPOINT_FOLDER", _env("RTU_SHAREPOINT_FOLDER", "RTU Report"))

COLUMNS = ["Branch", "Next Open Service", "Receive Consignment ID", "RCN Reference", "Consignor",
           "Consignee", "Number of Packages", "BKD", "ARV", "CTT", "PIC", "PUT", "In Warehouse",
           "Warehouse", "Closed", "Next Discharge Port", "Service Level", "Overs", "Booking Party",
           "Booked Under"]
DATE_COLS = {"Closed"}
WIDTHS = {"Branch": 8, "Next Open Service": 16, "Receive Consignment ID": 22.6, "RCN Reference": 14.9,
          "Consignor": 45, "Consignee": 45, "Number of Packages": 18, "In Warehouse": 12, "Warehouse": 30,
          "Closed": 15.6, "Next Discharge Port": 18, "Service Level": 22, "Booking Party": 45, "Booked Under": 12}
INWHS = {"ARV", "PUT", "PIC", "CTT", "STA", "FLO", "REC", "RCV"}

log = logging.getLogger("services_pending_report")
_cookies = http.cookiejar.CookieJar()
_opener  = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_cookies))

def _http(url, data=None, headers=None, method=None, timeout=180):
    return _opener.open(urllib.request.Request(url, data=data, headers=headers or {}, method=method), timeout=timeout)

def parse_dt_actual(s):
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})(?::(\d{2}))?(?:\.\d+)?(Z|[+-]\d{2}:?\d{2})?", s)
    if not m:
        return None
    y, mo, da, h, mi, se, off = m.groups()
    try:
        dt = datetime.datetime(int(y), int(mo), int(da), int(h), int(mi), int(se or 0))
    except ValueError:
        return None
    if off and off != "Z":
        sign = 1 if off[0] == "+" else -1
        dt -= datetime.timedelta(hours=sign * int(off[1:3]), minutes=sign * int(off[-2:]))
    dt += datetime.timedelta(hours=DISPLAY_TZ)
    return dt.replace(second=0, microsecond=0)

def to_instant(s):
    """Parse a DateTimeOffset to a sortable naive-UTC datetime (for 'next' ordering)."""
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})(?::(\d{2}))?(?:\.\d+)?(Z|[+-]\d{2}:?\d{2})?", s)
    if not m:
        return None
    y, mo, da, h, mi, se, off = m.groups()
    dt = datetime.datetime(int(y), int(mo), int(da), int(h), int(mi), int(se or 0))
    if off and off != "Z":
        sign = 1 if off[0] == "+" else -1
        dt -= datetime.timedelta(hours=sign * int(off[1:3]), minutes=sign * int(off[-2:]))
    return dt

class CargoWise:
    def __init__(self):
        self.auth = f"{CW_BASE}/Glow/auth/v2"; self.odata = f"{CW_BASE}/Glow/odata/{ODATA_MODEL}"
        self.user_key = None; self.branch_map = {}; self.dept_key = None

    def _post(self, path, body):
        r = _http(f"{self.auth}/{path}", data=json.dumps(body).encode(), method="POST",
                  headers={"Content-Type": "application/json", "Accept": "application/json",
                           "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop", "Origin": CW_BASE})
        return json.load(r)

    def _get(self, resource, params):
        url = f"{self.odata}/{resource}?{urllib.parse.urlencode(params)}"
        return json.load(_http(url, headers={"Accept": "application/json", "wtg-app": "Glow",
                                             "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop"}))

    def login(self):
        res = self._post("credential/claim/Staff",
                         {"userName": CW_USERNAME, "password": CW_PASSWORD, "setTokenCookie": True})
        if res.get("result") != 0:
            sys.exit(f"FATAL: CargoWise login failed (result={res.get('result')})")
        self.user_key = res["userKey"]
        ctx = self._post("credential/context/list",
                         {"logonProviderType": "Staff", "userKey": self.user_key, "useTokenCookie": True})
        for b in ctx.get("branchInfos", []):
            if b["code"].upper() in BRANCH_CODES:
                self.branch_map[b["code"].upper()] = b["key"]
        for d in ctx.get("departmentInfos", []):
            if d["code"].upper() == DEPARTMENT_CODE:
                self.dept_key = d["key"]
        missing = [c for c in BRANCH_CODES if c not in self.branch_map]
        if missing:
            sys.exit(f"FATAL: branch code(s) not found: {missing}")
        if not self.dept_key:
            sys.exit(f"FATAL: department {DEPARTMENT_CODE} not found")
        self._select(self.branch_map[BRANCH_CODES[0]])
        log.info("Authenticated as %s", res.get("userDisplayName"))

    def _select(self, branch_key):
        self._post("credential/context/select",
                   {"logonProviderType": "Staff", "userKey": self.user_key,
                    "branchKey": branch_key, "departmentKey": self.dept_key, "useAndSetTokenCookie": True})
        self._post("session/begin", {"tokenType": 1, "sessionType": "General", "useAndSetTokenCookie": True})

    def service_levels(self):
        try:
            res = self._get("RefServiceLevelInfos", {"$select": "RS_Code,RS_Description", "$top": "500"})
            return {r["RS_Code"]: r["RS_Description"] for r in res.get("value", [])}
        except Exception as e:
            log.warning("Could not load service levels (%s); using codes as-is", e)
            return {}

    def pull_branch(self, branch_key):
        self._select(branch_key)
        inwhs = " or ".join(f"p/WPS_Status eq '{s}'" for s in sorted(INWHS))
        # A service is booked AND no service has been completed (the grid's
        # "service completed date has no date" means NONE completed, not "one open").
        filt = (f"AdditionalServices/any(s: s/ES_BookedDateTimeOffset ne null) and "
                f"not AdditionalServices/any(s: s/ES_CompletedDateTimeOffset ne null) and "
                f"WhsItemPackageStates/any(p: {inwhs})")
        expand = ("Addresses($select=E2_AddressType;$expand=Address($expand=OrgHeader,Country)),"
                  "WhsItemPackageStates($select=WPS_Status;$expand=Package($select=KP_PackageQty)),"
                  "IntendedWarehouse($select=WW_WarehouseCode,WW_WarehouseName),"
                  "AdditionalServices($select=ES_ServiceCode,ES_BookedDateTimeOffset,ES_CompletedDateTimeOffset)")
        select = ("WRC_JobID,WRC_ConsignmentID,WRC_CompleteTime,"
                  "WRC_RL_NKNextDischargePort,WRC_RS_NKServiceLevel")
        rows, skip, page = [], 0, 50
        while True:
            params = {"$filter": filt, "$select": select, "$expand": expand,
                      "$top": str(page), "$skip": str(skip), "$orderby": "WRC_JobID"}
            for _ in range(4):
                try:
                    d = self._get("WhsItemReceiveConsignments", params); break
                except urllib.error.HTTPError as e:
                    if e.code == 401:
                        log.info("  session expired, re-authenticating"); self.login(); self._select(branch_key); continue
                    raise
            else:
                raise RuntimeError("repeated failures pulling consignments")
            batch = d.get("value", [])
            rows.extend(batch)
            if len(batch) < page:
                break
            skip += page
        return rows

def _addr(rec, atype):
    for a in rec.get("Addresses", []):
        if a.get("E2_AddressType") == atype:
            ad = a.get("Address") or {}; oh = ad.get("OrgHeader") or {}; ctry = ad.get("Country") or {}
            name = oh.get("OH_FullName") or ad.get("OA_CompanyNameOverride") or ""
            a1 = ad.get("OA_Address1") or ad.get("OA_Code") or ""
            a2 = ad.get("OA_Address2") or ""
            csz = " ".join(x for x in (ad.get("OA_City") or "", ad.get("OA_State") or "",
                                       ad.get("OA_PostCode") or "") if x)
            country = ctry.get("RN_Desc") or ad.get("OA_RN_NKCountryCode") or ""
            return ", ".join(x for x in (name, a1, a2, csz, country) if x)
    return ""

def next_open_service(rec):
    """The next open (booked, not completed) service. Ordered by booked date then by
    ES_ServiceId (first booked / first created). NOTE: for consignments with MULTIPLE
    open services, CargoWise's exact "next" ordering is not exposed in the OData feed
    (no sequence/priority field), so the pick may differ from the grid for those few
    rows; single-service consignments (the common case) are exact."""
    opens = [s for s in rec.get("AdditionalServices", []) if not s.get("ES_CompletedDateTimeOffset")]
    if not opens:
        return ""
    opens.sort(key=lambda s: (to_instant(s.get("ES_BookedDateTimeOffset")) is None,
                              to_instant(s.get("ES_BookedDateTimeOffset")) or datetime.datetime.max,
                              s.get("ES_ServiceId") or ""))
    return opens[0].get("ES_ServiceCode") or ""

def shape(branch, records, svc_map):
    out = []
    for r in records:
        ps = r.get("WhsItemPackageStates", [])
        def qty(statuses=None):
            return sum((p.get("Package") or {}).get("KP_PackageQty") or 0
                       for p in ps if statuses is None or p.get("WPS_Status") in statuses)
        wh = r.get("IntendedWarehouse") or {}
        wh_disp = f"{wh['WW_WarehouseCode']} - {wh.get('WW_WarehouseName') or ''}".strip(" -") if wh.get("WW_WarehouseCode") else ""
        sl = r.get("WRC_RS_NKServiceLevel") or ""
        sl = f"{sl} - {svc_map[sl]}" if sl in svc_map else sl
        out.append({
            "Branch": branch,
            "Next Open Service": next_open_service(r),
            "Receive Consignment ID": r.get("WRC_JobID"),
            "RCN Reference": r.get("WRC_ConsignmentID"),
            "Consignor": _addr(r, "CRG"),
            "Consignee": _addr(r, "CED"),
            "Number of Packages": qty(),
            "BKD": qty({"BKD"}), "ARV": qty({"ARV"}), "CTT": qty({"CTT"}),
            "PIC": qty({"PIC"}), "PUT": qty({"PUT"}), "In Warehouse": qty(INWHS),
            "Warehouse": wh_disp,
            "Closed": parse_dt_actual(r.get("WRC_CompleteTime")),
            "Next Discharge Port": r.get("WRC_RL_NKNextDischargePort") or "",
            "Service Level": sl,
            "Overs": 0,
            "Booking Party": _addr(r, "BKD"),
            "Booked Under": "RCN",
        })
    return out

def build_workbook(rows, path):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Receive Consignment"
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c) for c in COLUMNS])
    for ci, col in enumerate(COLUMNS, 1):
        L = get_column_letter(ci); ws.column_dimensions[L].width = WIDTHS.get(col, 11)
        if col in DATE_COLS:
            for cell in ws[L][1:]:
                cell.number_format = "dd-mmm-yy hh:mm"
    ws.freeze_panes = "A2"
    ws.add_table(Table(displayName="ReceiveConsignment",
                       ref=f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}",
                       tableStyleInfo=TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)))
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True); wb.save(path)

def upload(local_path):
    for k, v in {"AZURE_TENANT_ID": AZ_TENANT, "AZURE_CLIENT_ID": AZ_CLIENT, "AZURE_CLIENT_SECRET": AZ_SECRET,
                 "SHAREPOINT_HOSTNAME": SP_HOST, "SHAREPOINT_SITE_PATH": SP_SITE}.items():
        if not v:
            sys.exit(f"FATAL: UPLOAD=true but {k} is not set")
    graph = "https://graph.microsoft.com/v1.0"
    body = urllib.parse.urlencode({"client_id": AZ_CLIENT, "client_secret": AZ_SECRET,
                                   "scope": "https://graph.microsoft.com/.default",
                                   "grant_type": "client_credentials"}).encode()
    tok = json.load(_http(f"https://login.microsoftonline.com/{AZ_TENANT}/oauth2/v2.0/token", data=body,
                          method="POST", headers={"Content-Type": "application/x-www-form-urlencoded"}))["access_token"]
    H = {"Authorization": "Bearer " + tok}
    site = json.load(_http(f"{graph}/sites/{SP_HOST}:{SP_SITE}", headers=H))
    drive = json.load(_http(f"{graph}/sites/{site['id']}/drive", headers=H))
    dest = "/".join(p for p in (SP_FOLDER, os.path.basename(local_path)) if p)
    url = f"{graph}/drives/{drive['id']}/root:/{urllib.parse.quote(dest)}:/content"
    res = json.load(_http(url, data=open(local_path, "rb").read(), method="PUT",
                          headers={**H, "Content-Type": "application/octet-stream"}))
    log.info("Uploaded: %s (%s bytes)", res.get("webUrl"), res.get("size"))

def main():
    logging.basicConfig(level=logging.INFO, stream=sys.stdout, format="%(asctime)s %(levelname)s %(message)s")
    log.info("Services Pending (Freight On Hand) - branches=%s", BRANCH_CODES)
    cw = CargoWise(); cw.login()
    svc_map = cw.service_levels()
    all_rows = []
    for code in BRANCH_CODES:
        recs = cw.pull_branch(cw.branch_map[code])
        rows = shape(code, recs, svc_map)
        log.info("  %s: %d consignments", code, len(rows))
        all_rows.extend(rows)
    out_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE)
    build_workbook(all_rows, out_path)
    log.info("Workbook written: %s (%d rows)", out_path, len(all_rows))
    if DO_UPLOAD:
        upload(out_path)
    else:
        log.info("UPLOAD disabled; skipping SharePoint upload")
    log.info("Done.")

if __name__ == "__main__":
    main()
