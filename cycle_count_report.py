#!/usr/bin/env python3
"""
AWA Data Services - Cycle Count Report
======================================
Pulls Transit Cycle Count Locations from CargoWise TWD (WiseGrid) for the CON
branch and writes two Excel files matching the existing master format, then
(optionally) uploads them to SharePoint via Microsoft Graph (app-only).

  * "Cycle Count Data Full.xlsx"     - every cycle-count location record.
  * "Cycle Count Data 3 Months.xlsx" - sheet "Transit Cycle Count Location":
        sorted by End Time newest-first, then de-duplicated by Location
        (latest count per location). Plus a static "All Zones" reference sheet.

Config is environment-driven (see .env.example). No secrets are stored in code.
"""
import os, sys, json, re, logging, datetime, calendar, collections
import urllib.request, urllib.parse, urllib.error, http.cookiejar

# ---- tiny .env loader (no dependency) ----
def _load_dotenv(path=".env"):
    if os.path.exists(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())
_load_dotenv()

def _env(key, default=None, required=False):
    val = os.environ.get(key, default)
    if required and not val:
        sys.exit(f"FATAL: required environment variable {key} is not set")
    return val

CW_BASE      = _env("CW_BASE_URL", "https://www-isbint.wisegrid.net").rstrip("/")
CW_MODULE    = _env("CW_MODULE", "TWD")
ODATA_MODEL  = _env("CW_ODATA_MODEL", "TransitWarehouse")
CW_USERNAME  = _env("CW_USERNAME", required=True)
CW_PASSWORD  = _env("CW_PASSWORD", required=True)
BRANCH_CODE  = _env("CC_BRANCH_CODE", "CON").upper()
DEPT_CODE    = _env("CW_DEPARTMENT_CODE", "BRN").upper()
# CargoWise "Cycle Count Tasks" view filter: End Time was in the last N months (0 = no filter)
REPORT_MONTHS = int(_env("CC_REPORT_MONTHS", "12"))

OUTDIR        = _env("OUTPUT_DIR", "./output")
FULL_FILE     = _env("CC_FULL_FILENAME", "Cycle Count Data Full.xlsx")
DEDUP_FILE    = _env("CC_DEDUP_FILENAME", "Cycle Count Data 3 Months.xlsx")
DO_UPLOAD     = _env("UPLOAD", "true").lower() in ("1", "true", "yes")

AZ_TENANT  = _env("AZURE_TENANT_ID")
AZ_CLIENT  = _env("AZURE_CLIENT_ID")
AZ_SECRET  = _env("AZURE_CLIENT_SECRET")
SP_HOST    = _env("SHAREPOINT_HOSTNAME")
SP_SITE    = _env("SHAREPOINT_SITE_PATH")
SP_FOLDER  = _env("CC_SHAREPOINT_FOLDER", _env("SHAREPOINT_FOLDER", ""))

ALL_ZONES_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "cycle_count_all_zones.json")

STATUS_DESC = {"APP": "Approved", "REJ": "Rejected", "CMP": "Completed",
               "NST": "Not Started", "INP": "In Progress", "NEW": "New", "CAN": "Cancelled"}
COLUMNS = ["Cycle Count ID", "Warehouse", "Location", "Status", "Priority",
           "Assigned User", "Start Time", "End Time", "Is Recount", "Created Date"]
DATE_COLS = {"Start Time", "End Time", "Created Date"}

log = logging.getLogger("cycle_count_report")
_cookies = http.cookiejar.CookieJar()
_opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_cookies))

def _http(url, data=None, headers=None, method=None, timeout=180):
    return _opener.open(urllib.request.Request(url, data=data, headers=headers or {}, method=method), timeout=timeout)

def months_ago_iso(months):
    t = datetime.date.today()
    idx = t.month - 1 - months
    year = t.year + idx // 12
    month = idx % 12 + 1
    day = min(t.day, calendar.monthrange(year, month)[1])
    return f"{year:04d}-{month:02d}-{day:02d}T00:00:00Z"

def parse_dt(s):
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2}):?(\d{2})?", s)  # wall-clock
    if not m:
        return None
    y, mo, d, h, mi, se = m.groups()
    try:
        return datetime.datetime(int(y), int(mo), int(d), int(h), int(mi), int(se or 0))
    except ValueError:
        return None

def status_text(code):
    if not code:
        return ""
    return f"{code} - {STATUS_DESC[code]}" if code in STATUS_DESC else code

class CargoWise:
    def __init__(self):
        self.auth = f"{CW_BASE}/Glow/auth/v2"
        self.odata = f"{CW_BASE}/Glow/odata/{ODATA_MODEL}"
        self.user_key = None; self.branch_key = None; self.dept_key = None

    def _post(self, path, body):
        r = _http(f"{self.auth}/{path}", data=json.dumps(body).encode(), method="POST",
                  headers={"Content-Type": "application/json", "Accept": "application/json",
                           "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop", "Origin": CW_BASE})
        return json.load(r)

    def _get(self, resource, params):
        url = f"{self.odata}/{resource}?{urllib.parse.urlencode(params)}"
        return json.load(_http(url, headers={"Accept": "application/json", "wtg-app": "Glow",
                                             "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop"}))

    def login(self):
        res = self._post("credential/claim/Staff",
                         {"userName": CW_USERNAME, "password": CW_PASSWORD, "setTokenCookie": True})
        if res.get("result") != 0:
            sys.exit(f"FATAL: CargoWise login failed (result={res.get('result')})")
        self.user_key = res["userKey"]
        ctx = self._post("credential/context/list",
                         {"logonProviderType": "Staff", "userKey": self.user_key, "useTokenCookie": True})
        for b in ctx.get("branchInfos", []):
            if b["code"].upper() == BRANCH_CODE:
                self.branch_key = b["key"]
        for d in ctx.get("departmentInfos", []):
            if d["code"].upper() == DEPT_CODE:
                self.dept_key = d["key"]
        if not self.branch_key:
            sys.exit(f"FATAL: branch {BRANCH_CODE} not found for this user")
        if not self.dept_key:
            sys.exit(f"FATAL: department {DEPT_CODE} not found for this user")
        self._select()
        log.info("Authenticated as %s, branch %s", res.get("userDisplayName"), BRANCH_CODE)

    def _select(self):
        self._post("credential/context/select",
                   {"logonProviderType": "Staff", "userKey": self.user_key,
                    "branchKey": self.branch_key, "departmentKey": self.dept_key, "useAndSetTokenCookie": True})
        self._post("session/begin", {"tokenType": 1, "sessionType": "General", "useAndSetTokenCookie": True})

    def warehouse_label(self):
        try:
            res = self._get("WhsWarehouseInfos", {"$select": "WW_WarehouseCode,WW_WarehouseName", "$top": "1"})
            w = res["value"][0]
            return f"{w['WW_WarehouseCode']} - {w['WW_WarehouseName']}"
        except Exception as e:
            log.warning("Could not load warehouse label (%s)", e)
            return ""

    def pull_cycle_counts(self):
        rows, skip, page = [], 0, 50
        # Match the CargoWise view: End Time was in the last N months
        cutoff = months_ago_iso(REPORT_MONTHS) if REPORT_MONTHS > 0 else None
        base = {"$select": "WIC_JobID,WIC_Status,WIC_Priority,WIC_GS_NKAssignedTo,WIC_StartTime,"
                           "WIC_EndTime,WIC_WIC_RejectedCycleCount,WIC_SystemCreateTimeUtc",
                "$expand": "Location($select=WLV_LocationString)", "$orderby": "WIC_JobID"}
        if cutoff:
            base["$filter"] = f"WIC_EndTime ge {cutoff}"
        while True:
            params = {**base, "$top": str(page), "$skip": str(skip)}
            for _ in range(4):
                try:
                    d = self._get("WhsItemCycleCountLocations", params); break
                except urllib.error.HTTPError as e:
                    if e.code == 401:
                        log.info("  session expired, re-authenticating"); self.login(); continue
                    raise
            else:
                raise RuntimeError("repeated failures pulling cycle counts")
            batch = d.get("value", [])
            rows.extend(batch)
            if len(batch) < page:
                break
            skip += page
        return rows

def shape(records, warehouse):
    out = []
    for r in records:
        loc = (r.get("Location") or {}).get("WLV_LocationString") or ""
        out.append({
            "Cycle Count ID": r.get("WIC_JobID"),
            "Warehouse": warehouse,
            "Location": loc,
            "Status": status_text(r.get("WIC_Status")),
            "Priority": r.get("WIC_Priority"),
            "Assigned User": r.get("WIC_GS_NKAssignedTo"),
            "Start Time": parse_dt(r.get("WIC_StartTime")),
            "End Time": parse_dt(r.get("WIC_EndTime")),
            "Is Recount": "Y" if r.get("WIC_WIC_RejectedCycleCount") else "N",
            "Created Date": parse_dt(r.get("WIC_SystemCreateTimeUtc")),
        })
    return out

def dedup_latest_by_location(rows):
    # newest End Time first; keep first occurrence per Location (latest actual count)
    counted = [r for r in rows if r["End Time"] is not None]
    counted.sort(key=lambda r: r["End Time"], reverse=True)
    seen, result = set(), []
    for r in counted:
        loc = r["Location"]
        if loc in seen:
            continue
        seen.add(loc); result.append(r)
    return result

def _write_sheet(ws, columns, rows, table_name):
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    ws.append(columns)
    for r in rows:
        ws.append([r.get(c) for c in columns])
    for ci, col in enumerate(columns, 1):
        L = get_column_letter(ci)
        ws.column_dimensions[L].width = 24 if col in ("Cycle Count ID", "Warehouse", "Created Date",
                                                       "Start Time", "End Time") else 14
        if col in DATE_COLS:
            for cell in ws[L][1:]:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
    ws.freeze_panes = "A2"
    if rows:
        ws.add_table(Table(displayName=table_name,
                           ref=f"A1:{get_column_letter(len(columns))}{len(rows) + 1}",
                           tableStyleInfo=TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)))

def build_full(rows, path):
    import openpyxl
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _write_sheet(wb.create_sheet("Transit Cycle Count Location"), COLUMNS, rows, "CycleCountFull")
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True); wb.save(path)

def build_dedup(rows, path):
    import openpyxl
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _write_sheet(wb.create_sheet("Transit Cycle Count Location"), COLUMNS,
                 dedup_latest_by_location(rows), "CycleCountLatest")
    # static "All Zones" reference sheet (no header, as in the master file)
    if os.path.exists(ALL_ZONES_PATH):
        az = wb.create_sheet("All Zones")
        for row in json.load(open(ALL_ZONES_PATH)):
            az.append(row)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True); wb.save(path)

def upload(local_path):
    for k, v in {"AZURE_TENANT_ID": AZ_TENANT, "AZURE_CLIENT_ID": AZ_CLIENT, "AZURE_CLIENT_SECRET": AZ_SECRET,
                 "SHAREPOINT_HOSTNAME": SP_HOST, "SHAREPOINT_SITE_PATH": SP_SITE}.items():
        if not v:
            sys.exit(f"FATAL: UPLOAD=true but {k} is not set")
    graph = "https://graph.microsoft.com/v1.0"
    body = urllib.parse.urlencode({"client_id": AZ_CLIENT, "client_secret": AZ_SECRET,
                                   "scope": "https://graph.microsoft.com/.default",
                                   "grant_type": "client_credentials"}).encode()
    tok = json.load(_http(f"https://login.microsoftonline.com/{AZ_TENANT}/oauth2/v2.0/token", data=body,
                          method="POST", headers={"Content-Type": "application/x-www-form-urlencoded"}))["access_token"]
    H = {"Authorization": "Bearer " + tok}
    site = json.load(_http(f"{graph}/sites/{SP_HOST}:{SP_SITE}", headers=H))
    drive = json.load(_http(f"{graph}/sites/{site['id']}/drive", headers=H))
    dest = "/".join(p for p in (SP_FOLDER, os.path.basename(local_path)) if p)
    url = f"{graph}/drives/{drive['id']}/root:/{urllib.parse.quote(dest)}:/content"
    res = json.load(_http(url, data=open(local_path, "rb").read(), method="PUT",
                          headers={**H, "Content-Type": "application/octet-stream"}))
    log.info("Uploaded: %s (%s bytes)", res.get("webUrl"), res.get("size"))

def main():
    logging.basicConfig(level=logging.INFO, stream=sys.stdout, format="%(asctime)s %(levelname)s %(message)s")
    cw = CargoWise(); cw.login()
    warehouse = cw.warehouse_label()
    records = cw.pull_cycle_counts()
    rows = shape(records, warehouse)
    log.info("Pulled %d cycle-count records (%s)", len(rows), warehouse)

    full_path = os.path.join(OUTDIR, FULL_FILE)
    dedup_path = os.path.join(OUTDIR, DEDUP_FILE)
    build_full(rows, full_path)
    log.info("Wrote %s (%d rows)", full_path, len(rows))
    latest = dedup_latest_by_location(rows)
    build_dedup(rows, dedup_path)
    log.info("Wrote %s (%d unique locations)", dedup_path, len(latest))

    if DO_UPLOAD:
        upload(full_path); upload(dedup_path)
    else:
        log.info("UPLOAD disabled; skipping SharePoint upload")
    log.info("Done.")

if __name__ == "__main__":
    main()
