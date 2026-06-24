#!/usr/bin/env python3
"""
AWA Data Services - Open DTU's Report (CCC)
===========================================
The dispatch grid "CCC- DTU (Mika)" reproduced exactly: ALL dispatch transportation
units for the CCC branch (no active filter -> 1,304 units), with the same 13 columns
as the grid, then uploaded to SharePoint via the shared uploader.

Column sources (validated against the live CCC grid):
  Created time                    WDH_SystemCreateTimeUtc           (UTC-6 display)
  Dispatch transportation unit ID WDH_ReferenceNumber
  DTU reference                   WDH_VehicleReference
  Load list ID                    join of WDL_JobID over linked load lists
  Master bill                     join of load-list ReferenceNumbers CE_EntryType=='MAB'
  Transport company               TRA address
  Drivers name                    WDH_SignedBy
  FLO                             LoadedLoosePackagesCount
  Loaded handling units           TotalLoadedPackagesCount - LoadedLoosePackagesCount
  Remaining                       max(0, TotalPlannedPackagesCount - TotalLoadedPackagesCount)
  Staging location                join of load-list StagingLocation.WLV_LocationString
  Completion date                 WDH_LoadCompleteTime
  Status                          derived from load state (see status_of)

Config is environment-driven (see .env.example). No secrets are stored in code.
"""
import os, sys, json, re, logging, datetime
import urllib.request, urllib.parse, urllib.error, http.cookiejar

def _load_dotenv(path=".env"):
    if os.path.exists(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())
_load_dotenv()

def _env(key, default=None, required=False):
    val = os.environ.get(key, default)
    if required and not val:
        sys.exit(f"FATAL: required environment variable {key} is not set")
    return val

CW_BASE      = _env("CW_BASE_URL", "https://www-isbint.wisegrid.net").rstrip("/")
CW_MODULE    = _env("CW_MODULE", "TWD")
ODATA_MODEL  = _env("CW_ODATA_MODEL", "TransitWarehouse")
CW_USERNAME  = _env("CW_USERNAME", required=True)
CW_PASSWORD  = _env("CW_PASSWORD", required=True)
BRANCH_CODES = [b.strip().upper() for b in _env("ODTU_BRANCH_CODES", "CCC").split(",") if b.strip()]
DEPT_CODE    = _env("CW_DEPARTMENT_CODE", "BRN").upper()
DISPLAY_TZ_OFFSET = float(_env("CC_DISPLAY_TZ_OFFSET", "-6"))

OUTDIR       = _env("OUTPUT_DIR", "./output")
OUT_FILE     = _env("ODTU_FILENAME", "Open DTU's.xlsx")
DO_UPLOAD    = _env("UPLOAD", "true").lower() in ("1", "true", "yes")
SP_FOLDER    = _env("ODTU_SHAREPOINT_FOLDER", "DTU Report")

COLUMNS = ["Branch", "Created time", "Dispatch transportation unit ID", "DTU reference", "Load list ID",
           "Master bill", "Transport company", "FLO", "Loaded handling units",
           "Remaining", "Staging location", "Completion date", "Status", "Job number"]
DATE_COLS = {"Created time", "Completion date"}

log = logging.getLogger("open_dtu_report")
_cookies = http.cookiejar.CookieJar()
_opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_cookies))

def _http(url, data=None, headers=None, method=None, timeout=180):
    return _opener.open(urllib.request.Request(url, data=data, headers=headers or {}, method=method), timeout=timeout)

def parse_dt(s):
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})(?::(\d{2}))?(?:\.\d+)?(Z|[+-]\d{2}:?\d{2})?", s)
    if not m:
        return None
    y, mo, d, h, mi = (int(x) for x in m.groups()[:5])
    se = int(m.group(6) or 0); off = m.group(7)
    try:
        dt = datetime.datetime(y, mo, d, h, mi, se)
    except ValueError:
        return None
    if off and off != "Z":
        sign = 1 if off[0] == "+" else -1
        dt -= datetime.timedelta(hours=sign * int(off[1:3]), minutes=sign * int(off[-2:]))
    dt += datetime.timedelta(hours=DISPLAY_TZ_OFFSET)
    return dt.replace(second=0, microsecond=0)

class CargoWise:
    def __init__(self):
        self.auth = f"{CW_BASE}/Glow/auth/v2"; self.odata = f"{CW_BASE}/Glow/odata/{ODATA_MODEL}"
        self.user_key = None; self.branch_map = {}; self.dept_key = None

    def _post(self, path, body):
        r = _http(f"{self.auth}/{path}", data=json.dumps(body).encode(), method="POST",
                  headers={"Content-Type": "application/json", "Accept": "application/json",
                           "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop", "Origin": CW_BASE})
        return json.load(r)

    def _get(self, resource, params):
        url = f"{self.odata}/{resource}?{urllib.parse.urlencode(params)}"
        return json.load(_http(url, headers={"Accept": "application/json", "wtg-app": "Glow",
                                             "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop"}))

    def login(self):
        res = self._post("credential/claim/Staff",
                         {"userName": CW_USERNAME, "password": CW_PASSWORD, "setTokenCookie": True})
        if res.get("result") != 0:
            sys.exit(f"FATAL: CargoWise login failed (result={res.get('result')})")
        self.user_key = res["userKey"]
        ctx = self._post("credential/context/list",
                         {"logonProviderType": "Staff", "userKey": self.user_key, "useTokenCookie": True})
        for b in ctx.get("branchInfos", []):
            if b["code"].upper() in BRANCH_CODES:
                self.branch_map[b["code"].upper()] = b["key"]
        for d in ctx.get("departmentInfos", []):
            if d["code"].upper() == DEPT_CODE:
                self.dept_key = d["key"]
        missing = [c for c in BRANCH_CODES if c not in self.branch_map]
        if missing:
            sys.exit(f"FATAL: branch code(s) not found: {missing}")
        if not self.dept_key:
            sys.exit(f"FATAL: department {DEPT_CODE} not found")
        self._select(self.branch_map[BRANCH_CODES[0]])
        log.info("Authenticated as %s", res.get("userDisplayName"))

    def _select(self, branch_key):
        self._post("credential/context/select",
                   {"logonProviderType": "Staff", "userKey": self.user_key,
                    "branchKey": branch_key, "departmentKey": self.dept_key, "useAndSetTokenCookie": True})
        self._post("session/begin", {"tokenType": 1, "sessionType": "General", "useAndSetTokenCookie": True})

    def country_map(self):
        m, skip = {}, 0
        try:
            while True:
                d = self._get("RefCountryInfos", {"$select": "RN_Code,RN_Desc", "$top": "50", "$skip": str(skip)})
                v = d.get("value", [])
                m.update({r["RN_Code"]: r["RN_Desc"] for r in v})
                if len(v) < 50:
                    break
                skip += 50
        except Exception as e:
            log.warning("Could not load country map (%s)", e)
        return m

    def _paged(self, resource, params, branch_key, label):
        rows, skip, page = [], 0, 50
        while True:
            p = dict(params); p["$top"] = str(page); p["$skip"] = str(skip)
            for _ in range(4):
                try:
                    d = self._get(resource, p); break
                except urllib.error.HTTPError as e:
                    if e.code == 401:
                        log.info("  session expired, re-authenticating"); self.login(); self._select(branch_key); continue
                    raise
            else:
                raise RuntimeError(f"repeated failures pulling {label}")
            batch = d.get("value", [])
            rows.extend(batch)
            if len(batch) < page:
                break
            skip += page
            if skip % 1000 == 0:
                log.info("  ...%d", len(rows))
        return rows

    def pull_dtus(self, branch_key):
        self._select(branch_key)
        expand = ("Addresses($expand=Address($expand=OrgHeader)),"
                  "WhsItemDispatchLoadListDTUPivots($expand=TransitDispatchLoadList("
                  "$select=WDL_JobID;$expand=ReferenceNumbers($select=CE_EntryType,CE_EntryNum),"
                  "StagingLocation($select=WLV_LocationString)))")
        sel = "WDH_PK,WDH_ReferenceNumber,WDH_VehicleReference,WDH_SignedBy,WDH_LoadCompleteTime,WDH_SystemCreateTimeUtc"
        params = {"$expand": expand, "$select": sel, "$orderby": "WDH_ReferenceNumber"}
        return self._paged("WhsItemDispatchTransportationUnits", params, branch_key, "DTUs")

    def load_view_map(self, pks, branch_key):
        """Counts per DTU from TransitDispatchTransportationUnitLoadView, keyed by WDH_PK."""
        sel = ("WDH_PK,LoadedLoosePackagesCount,TotalLoadedPackagesCount,TotalPlannedPackagesCount")
        out = {}
        CH = 25
        for i in range(0, len(pks), CH):
            chunk = pks[i:i + CH]
            filt = " or ".join(f"WDH_PK eq {pk}" for pk in chunk)
            for _ in range(4):
                try:
                    d = self._get("TransitDispatchTransportationUnitLoadViews",
                                  {"$filter": filt, "$select": sel, "$top": str(CH)}); break
                except urllib.error.HTTPError as e:
                    if e.code == 401:
                        log.info("  session expired, re-authenticating"); self.login(); self._select(branch_key); continue
                    raise
            else:
                raise RuntimeError("repeated failures pulling load views")
            for r in d.get("value", []):
                out[r["WDH_PK"]] = r
        return out

def transport_company(dtu, cmap):
    for a in dtu.get("Addresses", []):
        if a.get("E2_AddressType") == "TRA":
            ad = a.get("Address") or {}; oh = ad.get("OrgHeader") or {}
            name = a.get("E2_CompanyName") or oh.get("OH_FullName") or ad.get("OA_CompanyNameOverride") or ""
            a1 = a.get("E2_Address1") or ad.get("OA_Address1") or ad.get("OA_Code") or ""
            a2 = a.get("E2_Address2") or ad.get("OA_Address2") or ""
            city = a.get("E2_City") or ad.get("OA_City") or ""
            state = a.get("E2_State") or ad.get("OA_State") or ""
            pc = a.get("E2_Postcode") or ad.get("OA_PostCode") or ""
            cc = a.get("E2_RN_NKCountryCode") or ad.get("OA_RN_NKCountryCode") or ""
            country = cmap.get(cc, cc)
            csz = " ".join(x for x in (city, state, pc) if x)
            return ", ".join(x for x in (name, a1, a2, csz, country) if x)
    return ""

def _join_distinct(values):
    seen, out = set(), []
    for v in values:
        if v and v not in seen:
            seen.add(v); out.append(v)
    return ", ".join(out)

def load_lists(dtu):
    return [(p.get("TransitDispatchLoadList") or {}) for p in dtu.get("WhsItemDispatchLoadListDTUPivots", [])]

def status_of(dtu, lv):
    """Derived dispatch status: Complete once load-complete is stamped, else
    Loading if anything has been loaded, else Open. (Best-effort - the grid's
    own 'Status' values were not visible to confirm exact wording.)"""
    if dtu.get("WDH_LoadCompleteTime"):
        return "Complete"
    if (lv or {}).get("TotalLoadedPackagesCount", 0):
        return "Loading"
    return "Open"

def shape(branch, records, lvmap, cmap):
    out = []
    for r in records:
        lv = lvmap.get(r["WDH_PK"], {})
        lls = load_lists(r)
        loose = lv.get("LoadedLoosePackagesCount", 0) or 0
        total_loaded = lv.get("TotalLoadedPackagesCount", 0) or 0
        total_planned = lv.get("TotalPlannedPackagesCount", 0) or 0
        mabs, jobs = [], []
        for ll in lls:
            for ce in ll.get("ReferenceNumbers", []):
                if ce.get("CE_EntryType") == "MAB":
                    mabs.append(ce.get("CE_EntryNum"))
                elif ce.get("CE_EntryType") == "FCO":
                    jobs.append(ce.get("CE_EntryNum"))
        out.append({
            "Branch": branch,
            "Created time": parse_dt(r.get("WDH_SystemCreateTimeUtc")),
            "Dispatch transportation unit ID": r.get("WDH_ReferenceNumber"),
            "DTU reference": r.get("WDH_VehicleReference"),
            "Load list ID": _join_distinct(ll.get("WDL_JobID") for ll in lls),
            "Master bill": _join_distinct(mabs),
            "Transport company": transport_company(r, cmap),
            "FLO": loose,
            "Loaded handling units": max(0, total_loaded - loose),
            "Remaining": 0 if r.get("WDH_LoadCompleteTime") else max(0, total_planned - total_loaded),
            "Staging location": _join_distinct((ll.get("StagingLocation") or {}).get("WLV_LocationString") for ll in lls),
            "Completion date": parse_dt(r.get("WDH_LoadCompleteTime")),
            "Status": status_of(r, lv),
            "Job number": _join_distinct(jobs),
        })
    return out

def build_workbook(rows, path):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    widths = {"Branch": 8, "Created time": 17, "Dispatch transportation unit ID": 26, "DTU reference": 20,
              "Load list ID": 28, "Master bill": 22, "Transport company": 40,
              "FLO": 8, "Loaded handling units": 18, "Remaining": 11, "Staging location": 16,
              "Completion date": 17, "Status": 12, "Job number": 16}
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Receive Transportation Unit"
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c) for c in COLUMNS])
    for ci, col in enumerate(COLUMNS, 1):
        L = get_column_letter(ci); ws.column_dimensions[L].width = widths.get(col, 12)
        if col in DATE_COLS:
            for cell in ws[L][1:]:
                cell.number_format = "dd-mmm-yy hh:mm"
    ws.freeze_panes = "A2"
    ws.add_table(Table(displayName="OpenDTU",
                       ref=f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}",
                       tableStyleInfo=TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)))
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True); wb.save(path)

def main():
    logging.basicConfig(level=logging.INFO, stream=sys.stdout, format="%(asctime)s %(levelname)s %(message)s")
    log.info("Open DTU's Report - branches=%s, all dispatch units (no filter)", BRANCH_CODES)
    cw = CargoWise(); cw.login()
    cmap = cw.country_map()
    all_rows = []
    for code in BRANCH_CODES:
        bkey = cw.branch_map[code]
        recs = cw.pull_dtus(bkey)
        lvmap = cw.load_view_map([r["WDH_PK"] for r in recs], bkey)
        rows = shape(code, recs, lvmap, cmap)
        log.info("  %s: %d DTUs", code, len(rows))
        all_rows.extend(rows)
    out_path = os.path.join(OUTDIR, OUT_FILE)
    build_workbook(all_rows, out_path)
    log.info("Workbook written: %s (%d rows)", out_path, len(all_rows))
    if DO_UPLOAD:
        import sp_upload; sp_upload.upload(out_path, SP_FOLDER)
    else:
        log.info("UPLOAD disabled; skipping SharePoint upload")
    log.info("Done.")

if __name__ == "__main__":
    main()
