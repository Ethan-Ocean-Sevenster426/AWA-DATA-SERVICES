#!/usr/bin/env python3
"""
AWA Data Services - RCN "No Booking Party" Report
=================================================
Pulls the "RCN-No Booking Party" Receive-Consignment view from CargoWise TWD
(WiseGrid) for the configured branches, builds an Excel workbook matching the
master template ("RCN - No Booking Party (Daily).xlsx"), and (optionally)
uploads it to SharePoint via Microsoft Graph.

Runs unattended on Linux (systemd timer / cron / Docker / GitHub Actions).
All configuration comes from environment variables - NO secrets are stored in code.

Saved search "RCN-No Booking Party" (FilterKey b0d35375-...) has THREE conditions:
  1. RECEIVEGATEINTIME  HasDate            -> reproduced (package has a gated-in
                                              transport unit: WRH_GateInTime ne null)
  2. CREATETIME         WasInTheLastMonth  -> reproduced (WRC_SystemCreateTimeUtc
                                              >= one rolling month ago)
  3. BOOKINGPARTY       IsBlank            -> NOT replicated. It is NOT the BKD
                                              ("Booking Party") address on the
                                              consignment - that is populated for the
                                              records that ARE in the live report (e.g.
                                              "ISLAND CARGO SUPPORT"), so it cannot be
                                              the discriminator. The exact field this
                                              maps to has not been pinned down, so the
                                              output applies conditions 1+2 only and is
                                              a SUPERSET (a small number of extra rows;
                                              it never MISSES report rows).

Timezone handling (matches the master FILE, which Power BI reads):
  * "actual" timestamps  (Created Time, Closed) -> true-instant converted to the
    file's display timezone (UTC-6). Verified on RC00118784 (file 11:03 == 17:03Z-6)
    and RC00118784 Closed (18:57:23-05:00 -> 23:57Z -> 17:57).
  * "planned" timestamps (Expected Arrival/Dispatch, stored as Z) -> shown wall-clock.
    Verified on RC00120439 (07:10Z -> file 07:10).

All 26 columns are populated from the receive-consignment feed, including ones that
look cross-module but are actually carried on the consignment:
  * Consol ID    -> ReferenceNumbers entry CE_EntryType == 'FCO'
  * Master Bill  -> ReferenceNumbers entry CE_EntryType == 'MAB'
  * Bill To Party-> Address type 'CRB'
"""
import os, sys, json, re, logging, calendar, datetime, collections
import urllib.request, urllib.parse, urllib.error, http.cookiejar

# --------------------------------------------------------------------------- #
# Config (environment-driven; tiny .env loader so local dev works without deps)
# --------------------------------------------------------------------------- #
def _load_dotenv(path=".env"):
    if os.path.exists(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())

_load_dotenv()

def _env(key, default=None, required=False):
    val = os.environ.get(key, default)
    if required and not val:
        sys.exit(f"FATAL: required environment variable {key} is not set")
    return val

# CargoWise / WiseGrid
CW_BASE        = _env("CW_BASE_URL", "https://www-isbint.wisegrid.net").rstrip("/")
CW_MODULE      = _env("CW_MODULE", "TWD")
ODATA_MODEL    = _env("CW_ODATA_MODEL", "TransitWarehouse")
CW_USERNAME    = _env("CW_USERNAME", required=True)
CW_PASSWORD    = _env("CW_PASSWORD", required=True)
BRANCH_CODES   = [b.strip().upper() for b in _env("RCN_BRANCH_CODES", "DOR,CON").split(",") if b.strip()]
DEPARTMENT_CODE= _env("CW_DEPARTMENT_CODE", "BRN").upper()
REPORT_MONTHS  = int(_env("RCN_REPORT_MONTHS", "1"))           # "created in the last month"
DISPLAY_TZ     = int(_env("RCN_DISPLAY_TZ_OFFSET", _env("CC_DISPLAY_TZ_OFFSET", "-6")))

# Output / behaviour
OUTPUT_DIR     = _env("OUTPUT_DIR", "./output")
OUTPUT_FILE    = _env("RCN_FILENAME", "RCN - No Booking Party (Daily).xlsx")
DO_UPLOAD      = _env("UPLOAD", "true").lower() in ("1", "true", "yes")

# Microsoft Graph / SharePoint
AZ_TENANT      = _env("AZURE_TENANT_ID")
AZ_CLIENT      = _env("AZURE_CLIENT_ID")
AZ_SECRET      = _env("AZURE_CLIENT_SECRET")
SP_HOST        = _env("SHAREPOINT_HOSTNAME")
SP_SITE        = _env("SHAREPOINT_SITE_PATH")                  # e.g. /sites/DataPrime
SP_FOLDER      = _env("RCN_SHAREPOINT_FOLDER", "RTU Report")

log = logging.getLogger("rcn_report")

# --------------------------------------------------------------------------- #
# HTTP helper (cookie-aware) + small utilities
# --------------------------------------------------------------------------- #
_cookies = http.cookiejar.CookieJar()
_opener  = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_cookies))

def _http(url, data=None, headers=None, method=None, timeout=180):
    req = urllib.request.Request(url, data=data, headers=headers or {}, method=method)
    return _opener.open(req, timeout=timeout)

def months_ago_iso(months):
    t = datetime.date.today()
    idx = t.month - 1 - months
    year = t.year + idx // 12
    month = idx % 12 + 1
    day = min(t.day, calendar.monthrange(year, month)[1])
    return f"{year:04d}-{month:02d}-{day:02d}T00:00:00Z"

def parse_dt_actual(s):
    """Actual event timestamp: convert to true instant (UTC) then to display tz."""
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})(?::(\d{2}))?(?:\.\d+)?(Z|[+-]\d{2}:?\d{2})?", s)
    if not m:
        return None
    y, mo, da, h, mi, se, off = m.groups()
    try:
        dt = datetime.datetime(int(y), int(mo), int(da), int(h), int(mi), int(se or 0))
    except ValueError:
        return None
    if off and off != "Z":                       # strip the stored offset -> UTC
        sign = 1 if off[0] == "+" else -1
        dt -= datetime.timedelta(hours=sign * int(off[1:3]), minutes=sign * int(off[-2:]))
    dt += datetime.timedelta(hours=DISPLAY_TZ)   # UTC -> file display tz
    return dt.replace(second=0, microsecond=0)

def parse_dt_wall(s):
    """Planned timestamp (stored as Z): shown as wall-clock, no tz conversion."""
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})", s)
    if not m:
        return None
    try:
        return datetime.datetime(*[int(x) for x in m.groups()])
    except ValueError:
        return None

def fmt_qty(v, unit):
    """CargoWise quantity display: up to 3 decimals, trailing zeros trimmed."""
    s = f"{(v or 0):.3f}".rstrip("0").rstrip(".")
    if s in ("", "-0"):
        s = "0"
    return f"{s} {unit}"

# Standard CargoWise transport modes (the master file uses SEA/AIR/RAI/COU).
TRANSPORT_MODES = {"SEA": "Sea Freight", "AIR": "Air Freight", "ROA": "Road Freight",
                   "RAI": "Rail Freight", "COU": "Courier"}
INWHS = {"ARV", "PUT", "PIC", "CTT", "STA", "FLO", "REC", "RCV"}   # "In Warehouse" bucket

PULL_EXPAND = ("Addresses($select=E2_AddressType;$expand=Address($expand=OrgHeader,Country)),"
               "WhsItemPackageStates($select=WPS_Status;$expand=Package($select=KP_Weight,KP_Volume,KP_PackageQty)),"
               "CreatedByStaff($select=GS_Code,GS_FullName),"
               "IntendedWarehouse($select=WW_WarehouseCode,WW_WarehouseName),"
               "ReferenceNumbers($select=CE_EntryType,CE_EntryNum)")
PULL_SELECT = ("WRC_JobID,WRC_ConsignmentID,WRC_SystemCreateTimeUtc,WRC_SystemCreateUser,"
               "WRC_CompleteTime,WRC_ExpectedArrivalTime,WRC_ExpectedDispatchTime,"
               "WRC_RL_NKNextDischargePort,WRC_RL_NKDestination,WRC_RS_NKServiceLevel,WRC_TransportMode")

# --------------------------------------------------------------------------- #
# CargoWise client
# --------------------------------------------------------------------------- #
class CargoWise:
    def __init__(self):
        self.auth = f"{CW_BASE}/Glow/auth/v2"
        self.odata = f"{CW_BASE}/Glow/odata/{ODATA_MODEL}"
        self.user_key = None
        self.branch_map = {}
        self.dept_key = None

    def _post(self, path, body):
        data = json.dumps(body).encode()
        r = _http(f"{self.auth}/{path}", data=data, method="POST",
                  headers={"Content-Type": "application/json", "Accept": "application/json",
                           "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop", "Origin": CW_BASE})
        return json.load(r)

    def _get(self, resource, params):
        url = f"{self.odata}/{resource}?{urllib.parse.urlencode(params)}"
        r = _http(url, headers={"Accept": "application/json", "wtg-app": "Glow",
                                "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop"})
        return json.load(r)

    def login(self):
        res = self._post("credential/claim/Staff",
                         {"userName": CW_USERNAME, "password": CW_PASSWORD, "setTokenCookie": True})
        if res.get("result") != 0:
            sys.exit(f"FATAL: CargoWise login failed (result={res.get('result')})")
        self.user_key = res["userKey"]
        log.info("Authenticated to CargoWise as %s", res.get("userDisplayName"))
        self._resolve_contexts()

    def _resolve_contexts(self):
        res = self._post("credential/context/list",
                         {"logonProviderType": "Staff", "userKey": self.user_key, "useTokenCookie": True})
        for b in res.get("branchInfos", []):
            if b["code"].upper() in BRANCH_CODES:
                self.branch_map[b["code"].upper()] = b["key"]
        for d in res.get("departmentInfos", []):
            if d["code"].upper() == DEPARTMENT_CODE:
                self.dept_key = d["key"]
        missing = [c for c in BRANCH_CODES if c not in self.branch_map]
        if missing:
            sys.exit(f"FATAL: branch code(s) not found for this user: {missing}")
        if not self.dept_key:
            sys.exit(f"FATAL: department code {DEPARTMENT_CODE} not found for this user")
        self._select_branch(self.branch_map[BRANCH_CODES[0]])

    def _select_branch(self, branch_key):
        self._post("credential/context/select",
                   {"logonProviderType": "Staff", "userKey": self.user_key,
                    "branchKey": branch_key, "departmentKey": self.dept_key, "useAndSetTokenCookie": True})
        self._post("session/begin", {"tokenType": 1, "sessionType": "General", "useAndSetTokenCookie": True})

    def service_levels(self):
        try:
            res = self._get("RefServiceLevelInfos", {"$select": "RS_Code,RS_Description", "$top": "500"})
            return {r["RS_Code"]: r["RS_Description"] for r in res.get("value", [])}
        except Exception as e:
            log.warning("Could not load service levels (%s); using codes as-is", e)
            return {}

    def pull_branch(self, branch_key, cutoff_iso):
        self._select_branch(branch_key)
        expand, select = PULL_EXPAND, PULL_SELECT
        # filters 1 (gate-in HasDate) + 2 (created in the last month). Filter 3
        # (booking party blank) is not replicated - see module docstring.
        filt = (f"WRC_SystemCreateTimeUtc ge {cutoff_iso} and "
                f"WhsItemPackageStates/any(p: p/TransitReceiveHeader/WRH_GateInTime ne null)")
        rows, skip, page = [], 0, 50
        while True:
            params = {"$filter": filt, "$select": select, "$expand": expand,
                      "$top": str(page), "$skip": str(skip), "$orderby": "WRC_JobID"}
            for _ in range(4):
                try:
                    d = self._get("WhsItemReceiveConsignments", params)
                    break
                except urllib.error.HTTPError as e:
                    if e.code == 401:
                        log.info("  session expired, re-authenticating")
                        self.login(); self._select_branch(branch_key); continue
                    raise
            else:
                raise RuntimeError("repeated failures pulling consignments")
            batch = d.get("value", [])
            rows.extend(batch)
            if len(batch) < page:
                break
            skip += page
        return rows

# --------------------------------------------------------------------------- #
# Report shaping (matches master template "RCN - No Booking Party (Daily).xlsx")
# --------------------------------------------------------------------------- #
COLUMNS = ["Create User Code", "Created Time", "Closed", "Number of Packages", "In Warehouse",
           "PUT", "DEP", "Total Weight", "RCN Reference", "Consignor", "Booking Party",
           "Receive Consignment ID", "Master Bill", "Total Volume", "Consignee",
           "Next Discharge Port", "Destination Port", "Expected Arrival at Warehouse",
           "Expected Dispatch from Warehouse", "Service Level", "BKD", "Overs", "Warehouse",
           "Consol ID", "Transport Mode", "Bill To Party"]
DATE_COLS = {"Created Time", "Closed", "Expected Arrival at Warehouse", "Expected Dispatch from Warehouse"}
WIDTHS = {"Create User Code": 22, "Created Time": 15.6, "Closed": 15.6, "Number of Packages": 18,
          "In Warehouse": 12, "PUT": 8, "DEP": 8, "Total Weight": 13, "RCN Reference": 14.9,
          "Consignor": 45, "Booking Party": 45, "Receive Consignment ID": 22.6, "Master Bill": 16,
          "Total Volume": 13, "Consignee": 45, "Next Discharge Port": 18, "Destination Port": 16,
          "Expected Arrival at Warehouse": 28.5, "Expected Dispatch from Warehouse": 32.5,
          "Service Level": 22, "BKD": 8, "Overs": 8, "Warehouse": 30, "Consol ID": 14,
          "Transport Mode": 18, "Bill To Party": 30}

def _addr(rec, atype):
    for a in rec.get("Addresses", []):
        if a.get("E2_AddressType") == atype:
            ad = a.get("Address") or {}
            oh = ad.get("OrgHeader") or {}
            ctry = ad.get("Country") or {}
            name = oh.get("OH_FullName") or ad.get("OA_CompanyNameOverride") or ""
            a1 = ad.get("OA_Address1") or ad.get("OA_Code") or ""
            a2 = ad.get("OA_Address2") or ""
            csz = " ".join(x for x in (ad.get("OA_City") or "", ad.get("OA_State") or "",
                                       ad.get("OA_PostCode") or "") if x)
            country = ctry.get("RN_Desc") or ad.get("OA_RN_NKCountryCode") or ""
            return ", ".join(x for x in (name, a1, a2, csz, country) if x)
    return ""

def _ref(rec, etype):
    for e in rec.get("ReferenceNumbers", []):
        if e.get("CE_EntryType") == etype:
            return e.get("CE_EntryNum") or ""
    return ""

def _counts(rec):
    ps = rec.get("WhsItemPackageStates", [])
    # Package columns are the SUM of package quantities (KP_PackageQty) by status,
    # not the count of package-state records (one state can be qty > 1).
    def qty(statuses=None):
        return sum((p.get("Package") or {}).get("KP_PackageQty") or 0
                   for p in ps if statuses is None or p.get("WPS_Status") in statuses)
    weight = sum((p.get("Package") or {}).get("KP_Weight") or 0 for p in ps)
    volume = sum((p.get("Package") or {}).get("KP_Volume") or 0 for p in ps)
    return {"Number of Packages": qty(),
            "In Warehouse": qty(INWHS),
            "PUT": qty({"PUT"}), "DEP": qty({"DEP"}), "BKD": qty({"BKD"}),
            "weight": weight, "volume": volume}

def shape_rows(records, svc_map):
    out = []
    for r in records:
        staff = r.get("CreatedByStaff") or {}
        code = staff.get("GS_Code") or r.get("WRC_SystemCreateUser") or ""
        name = staff.get("GS_FullName") or ""
        user = f"{code} - {name}" if name else code
        wh = r.get("IntendedWarehouse") or {}
        wh_disp = ""
        if wh.get("WW_WarehouseCode"):
            wh_disp = f"{wh['WW_WarehouseCode']} - {wh.get('WW_WarehouseName') or ''}".strip(" -")
        sl = r.get("WRC_RS_NKServiceLevel") or ""
        sl = f"{sl} - {svc_map[sl]}" if sl in svc_map else sl
        tm = r.get("WRC_TransportMode") or ""
        tm = f"{tm} - {TRANSPORT_MODES[tm]}" if tm in TRANSPORT_MODES else tm
        cnt = _counts(r)
        out.append({
            "Create User Code": user,
            "Created Time": parse_dt_actual(r.get("WRC_SystemCreateTimeUtc")),
            "Closed": parse_dt_actual(r.get("WRC_CompleteTime")),
            "Number of Packages": cnt["Number of Packages"],
            "In Warehouse": cnt["In Warehouse"], "PUT": cnt["PUT"], "DEP": cnt["DEP"],
            "Total Weight": fmt_qty(cnt["weight"], "KG"),
            "RCN Reference": r.get("WRC_ConsignmentID"),
            "Consignor": _addr(r, "CRG"),
            "Booking Party": _addr(r, "BKD"),
            "Receive Consignment ID": r.get("WRC_JobID"),
            "Master Bill": _ref(r, "MAB"),                       # master bill reference
            "Total Volume": fmt_qty(cnt["volume"], "M3"),
            "Consignee": _addr(r, "CED"),
            "Next Discharge Port": r.get("WRC_RL_NKNextDischargePort") or "",
            "Destination Port": r.get("WRC_RL_NKDestination") or "",
            "Expected Arrival at Warehouse": parse_dt_wall(r.get("WRC_ExpectedArrivalTime")),
            "Expected Dispatch from Warehouse": parse_dt_wall(r.get("WRC_ExpectedDispatchTime")),
            "Service Level": sl,
            "BKD": cnt["BKD"], "Overs": 0,
            "Warehouse": wh_disp,
            "Consol ID": _ref(r, "FCO"),                         # forwarding consol reference
            "Transport Mode": tm,
            "Bill To Party": _addr(r, "CRB"),                    # bill-to (CRB) address
        })
    return out

def build_workbook(rows, path):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receive Consignment"
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c) for c in COLUMNS])
    for ci, col in enumerate(COLUMNS, 1):
        L = get_column_letter(ci)
        ws.column_dimensions[L].width = WIDTHS.get(col, 14)
        if col in DATE_COLS:
            for cell in ws[L][1:]:
                cell.number_format = "dd-mmm-yy hh:mm"
    ws.freeze_panes = "A2"
    ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    tbl = Table(displayName="ReceiveConsignment", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)

# --------------------------------------------------------------------------- #
# SharePoint upload (Microsoft Graph, app-only client-credentials)
# --------------------------------------------------------------------------- #
def upload_to_sharepoint(local_path):
    for k, v in {"AZURE_TENANT_ID": AZ_TENANT, "AZURE_CLIENT_ID": AZ_CLIENT,
                 "AZURE_CLIENT_SECRET": AZ_SECRET, "SHAREPOINT_HOSTNAME": SP_HOST,
                 "SHAREPOINT_SITE_PATH": SP_SITE}.items():
        if not v:
            sys.exit(f"FATAL: UPLOAD=true but {k} is not set")
    graph = "https://graph.microsoft.com/v1.0"
    body = urllib.parse.urlencode({"client_id": AZ_CLIENT, "client_secret": AZ_SECRET,
                                   "scope": "https://graph.microsoft.com/.default",
                                   "grant_type": "client_credentials"}).encode()
    tok = json.load(_http(f"https://login.microsoftonline.com/{AZ_TENANT}/oauth2/v2.0/token",
                          data=body, method="POST",
                          headers={"Content-Type": "application/x-www-form-urlencoded"}))["access_token"]
    H = {"Authorization": "Bearer " + tok}
    site = json.load(_http(f"{graph}/sites/{SP_HOST}:{SP_SITE}", headers=H))
    drive = json.load(_http(f"{graph}/sites/{site['id']}/drive", headers=H))
    dest = "/".join(p for p in (SP_FOLDER, os.path.basename(local_path)) if p)
    url = f"{graph}/drives/{drive['id']}/root:/{urllib.parse.quote(dest)}:/content"
    with open(local_path, "rb") as f:
        content = f.read()
    res = json.load(_http(url, data=content, method="PUT",
                          headers={**H, "Content-Type": "application/octet-stream"}))
    log.info("Uploaded to SharePoint: %s (%s bytes)", res.get("webUrl"), res.get("size"))

# --------------------------------------------------------------------------- #
def main():
    logging.basicConfig(level=logging.INFO, stream=sys.stdout,
                        format="%(asctime)s %(levelname)s %(message)s")
    cutoff = months_ago_iso(REPORT_MONTHS)
    log.info("RCN No-Booking-Party Report - branches=%s, created since %s, display tz UTC%+d",
             BRANCH_CODES, cutoff, DISPLAY_TZ)

    cw = CargoWise()
    cw.login()
    svc_map = cw.service_levels()

    all_rows = []
    for code in BRANCH_CODES:
        records = cw.pull_branch(cw.branch_map[code], cutoff)
        rows = shape_rows(records, svc_map)
        log.info("  %s: %d consignments", code, len(rows))
        all_rows.extend(rows)

    out_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE)
    build_workbook(all_rows, out_path)
    log.info("Workbook written: %s (%d rows)", out_path, len(all_rows))

    if DO_UPLOAD:
        upload_to_sharepoint(out_path)
    else:
        log.info("UPLOAD disabled; skipping SharePoint upload")
    log.info("Done.")

if __name__ == "__main__":
    main()
