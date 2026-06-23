#!/usr/bin/env python3
"""
AWA Data Services - PKG "No Consignment" Report
===============================================
Pulls the "PKG-No Consignment" saved search (packages received in the last 14 days
that are not yet linked to a receive consignment) from CargoWise TWD (WiseGrid)
for both branches (DOR + CON), and writes a single "Transit Package" sheet matching
the master "PKG - No Consignment (Daily).xlsx", then (optionally) uploads to
SharePoint via Microsoft Graph.

Saved search "PKG-No Consignment" (module IEntityInfo_IWhsItemPackageState):
  * RECEIVEGATEINTIME     WasInTheLast14Days -> RTU gate-in within the last 14 days
  * RECEIPTCONSIGNMENTID  IsBlank            -> package not linked to a consignment
Both conditions reproduced exactly (validated: returns the same 4 CON rows the live
grid shows).

Because these packages have NO consignment (and therefore no dispatch / shipment /
ASN / load-list links), the consignment- and dispatch-derived columns are blank by
nature - this matches the master file. Populated columns come from the package
itself, its RTU (gate-in/out, unload), location and status.

Timezone: gate-in / unload / gate-out are actual events -> true-instant converted to
the file display tz (UTC-6), matching the master and the other reports.

Config is environment-driven (see .env.example). No secrets are stored in code.
"""
import os, sys, json, re, logging, datetime
import urllib.request, urllib.parse, urllib.error, http.cookiejar

def _load_dotenv(path=".env"):
    if os.path.exists(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())
_load_dotenv()

def _env(key, default=None, required=False):
    val = os.environ.get(key, default)
    if required and not val:
        sys.exit(f"FATAL: required environment variable {key} is not set")
    return val

CW_BASE      = _env("CW_BASE_URL", "https://www-isbint.wisegrid.net").rstrip("/")
CW_MODULE    = _env("CW_MODULE", "TWD")
ODATA_MODEL  = _env("CW_ODATA_MODEL", "TransitWarehouse")
CW_USERNAME  = _env("CW_USERNAME", required=True)
CW_PASSWORD  = _env("CW_PASSWORD", required=True)
BRANCH_CODES = [b.strip().upper() for b in _env("PKG_BRANCH_CODES", "DOR,CON").split(",") if b.strip()]
DEPT_CODE    = _env("CW_DEPARTMENT_CODE", "BRN").upper()
GATE_DAYS    = int(_env("PKG_GATE_DAYS", "14"))                 # "gate-in in the last N days"
DISPLAY_TZ_OFFSET = float(_env("CC_DISPLAY_TZ_OFFSET", "-6"))   # CargoWise fixed display offset

OUTDIR       = _env("OUTPUT_DIR", "./output")
OUT_FILE     = _env("PKG_FILENAME", "PKG - No Consignment (Daily).xlsx")
DO_UPLOAD    = _env("UPLOAD", "true").lower() in ("1", "true", "yes")

AZ_TENANT  = _env("AZURE_TENANT_ID"); AZ_CLIENT = _env("AZURE_CLIENT_ID"); AZ_SECRET = _env("AZURE_CLIENT_SECRET")
SP_HOST    = _env("SHAREPOINT_HOSTNAME"); SP_SITE = _env("SHAREPOINT_SITE_PATH")
SP_FOLDER  = _env("PKG_SHAREPOINT_FOLDER", _env("RTU_SHAREPOINT_FOLDER", "RTU Report"))

COLUMNS = ["Branch", "Receive Transportation Unit ID", "Receive Gate in Time", "RTU Unload Complete",
           "Receive Gate out Time", "RCN Reference", "Package ID", "Quantity", "Package Type",
           "Weight", "Length", "Width", "Height", "Volume", "Location", "Status", "DTU Load Complete",
           "Dispatch Gate in Time", "Dispatch Gate out Time", "DTU External Reference", "Handling Unit ID",
           "Expected Arrival at Warehouse", "Expected Dispatch from Warehouse", "Booking Party",
           "Consignor", "Consignee", "Commodity", "Description", "Dispatch Consol ID",
           "Dispatch Master Bill Number", "Dispatch House Bill Number", "Dispatch Shipment ID",
           "Load List ID", "Receive Consignment ID", "Receive Shipment ID", "Receive Master Bill Number",
           "ASN Closed", "DCN Closed", "Load List Closed", "Marks and Numbers", "Is Ready to Stage",
           "RCN Closed"]
DATE_COLS = {"Receive Gate in Time", "RTU Unload Complete", "Receive Gate out Time", "DTU Load Complete",
             "Dispatch Gate in Time", "Dispatch Gate out Time", "Expected Arrival at Warehouse",
             "Expected Dispatch from Warehouse", "ASN Closed", "DCN Closed", "Load List Closed", "RCN Closed"}

# WPS_Status -> description. No ref entity is exposed; the standard transit-warehouse
# statuses are mapped here, and any code not listed falls back to the bare code.
STATUS_DESC = {"ARV": "Arrived", "PUT": "Put Away", "PIC": "Picked", "STA": "Staged",
               "LOA": "Loaded", "DEP": "Departed", "BKD": "Booked"}

log = logging.getLogger("pkg_report")
_cookies = http.cookiejar.CookieJar()
_opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_cookies))

def _http(url, data=None, headers=None, method=None, timeout=180):
    return _opener.open(urllib.request.Request(url, data=data, headers=headers or {}, method=method), timeout=timeout)

def days_ago_boundary(days):
    """Start of the day 'days' ago, in display tz, as the equivalent UTC instant."""
    d = datetime.date.today() - datetime.timedelta(days=days)
    boundary_h = int(round(-DISPLAY_TZ_OFFSET))            # -6 -> 06:00Z
    return f"{d.isoformat()}T{boundary_h:02d}:00:00Z"

def parse_dt(s):
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})(?::(\d{2}))?(?:\.\d+)?(Z|[+-]\d{2}:?\d{2})?", s)
    if not m:
        return None
    y, mo, d, h, mi = (int(x) for x in m.groups()[:5])
    se = int(m.group(6) or 0); off = m.group(7)
    try:
        dt = datetime.datetime(y, mo, d, h, mi, se)
    except ValueError:
        return None
    if off and off != "Z":
        sign = 1 if off[0] == "+" else -1
        dt -= datetime.timedelta(hours=sign * int(off[1:3]), minutes=sign * int(off[-2:]))
    dt += datetime.timedelta(hours=DISPLAY_TZ_OFFSET)
    return dt.replace(second=0, microsecond=0)

def fmt_qty(v, unit):
    """CargoWise quantity display: up to 3 decimals, trailing zeros trimmed, + unit."""
    s = f"{(v or 0):.3f}".rstrip("0").rstrip(".")
    if s in ("", "-0"):
        s = "0"
    return f"{s} {unit}".strip() if unit else s

class CargoWise:
    def __init__(self):
        self.auth = f"{CW_BASE}/Glow/auth/v2"; self.odata = f"{CW_BASE}/Glow/odata/{ODATA_MODEL}"
        self.user_key = None; self.branch_map = {}; self.dept_key = None

    def _post(self, path, body):
        r = _http(f"{self.auth}/{path}", data=json.dumps(body).encode(), method="POST",
                  headers={"Content-Type": "application/json", "Accept": "application/json",
                           "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop", "Origin": CW_BASE})
        return json.load(r)

    def _get(self, resource, params):
        url = f"{self.odata}/{resource}?{urllib.parse.urlencode(params)}"
        return json.load(_http(url, headers={"Accept": "application/json", "wtg-app": "Glow",
                                             "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop"}))

    def login(self):
        res = self._post("credential/claim/Staff",
                         {"userName": CW_USERNAME, "password": CW_PASSWORD, "setTokenCookie": True})
        if res.get("result") != 0:
            sys.exit(f"FATAL: CargoWise login failed (result={res.get('result')})")
        self.user_key = res["userKey"]
        ctx = self._post("credential/context/list",
                         {"logonProviderType": "Staff", "userKey": self.user_key, "useTokenCookie": True})
        for b in ctx.get("branchInfos", []):
            if b["code"].upper() in BRANCH_CODES:
                self.branch_map[b["code"].upper()] = b["key"]
        for d in ctx.get("departmentInfos", []):
            if d["code"].upper() == DEPT_CODE:
                self.dept_key = d["key"]
        missing = [c for c in BRANCH_CODES if c not in self.branch_map]
        if missing:
            sys.exit(f"FATAL: branch code(s) not found: {missing}")
        if not self.dept_key:
            sys.exit(f"FATAL: department {DEPT_CODE} not found")
        self._select(self.branch_map[BRANCH_CODES[0]])
        log.info("Authenticated as %s", res.get("userDisplayName"))

    def _select(self, branch_key):
        self._post("credential/context/select",
                   {"logonProviderType": "Staff", "userKey": self.user_key,
                    "branchKey": branch_key, "departmentKey": self.dept_key, "useAndSetTokenCookie": True})
        self._post("session/begin", {"tokenType": 1, "sessionType": "General", "useAndSetTokenCookie": True})

    def packtype_map(self):
        try:
            d = self._get("RefPackTypeInfos", {"$select": "F3_Code,F3_Description", "$top": "500"})
            return {r["F3_Code"]: r["F3_Description"] for r in d.get("value", [])}
        except Exception as e:
            log.warning("Could not load pack types (%s); using codes as-is", e)
            return {}

    def pull_packages(self, branch_key, cutoff_iso):
        self._select(branch_key)
        filt = (f"WPS_WRC_TransitReceiveConsignment eq null and "
                f"TransitReceiveHeader/WRH_GateInTime ge {cutoff_iso}")
        expand = ("Package($expand=PackageHeader($select=KPH_PackageID)),"
                  "TransitReceiveHeader($select=WRH_ReferenceNumber,WRH_GateInTime,WRH_GateOutTime,WRH_UnloadCompleteTime),"
                  "LastLocation($select=WLV_LocationString)")
        sel = "WPS_PK,WPS_Status"
        rows, skip, page = [], 0, 50
        while True:
            params = {"$filter": filt, "$expand": expand, "$select": sel,
                      "$top": str(page), "$skip": str(skip), "$orderby": "WPS_PK"}
            for _ in range(4):
                try:
                    d = self._get("WhsItemPackageStates", params); break
                except urllib.error.HTTPError as e:
                    if e.code == 401:
                        log.info("  session expired, re-authenticating"); self.login(); self._select(branch_key); continue
                    raise
            else:
                raise RuntimeError("repeated failures pulling packages")
            batch = d.get("value", [])
            rows.extend(batch)
            if len(batch) < page:
                break
            skip += page
            if skip % 1000 == 0:
                log.info("  ...%d", len(rows))
        return rows

def shape(branch, records, ptmap):
    out = []
    for r in records:
        p = r.get("Package") or {}
        kph = p.get("PackageHeader") or {}
        rtu = r.get("TransitReceiveHeader") or {}
        loc = r.get("LastLocation") or {}
        pt = p.get("KP_F3_NKPackType") or ""
        pt_disp = f"{pt} - {ptmap[pt]}" if pt in ptmap else pt
        st = r.get("WPS_Status") or ""
        st_disp = f"{st} - {STATUS_DESC[st]}" if st in STATUS_DESC else st
        row = {c: None for c in COLUMNS}                       # blanks by default
        row.update({
            "Branch": branch,
            "Receive Transportation Unit ID": rtu.get("WRH_ReferenceNumber"),
            "Receive Gate in Time": parse_dt(rtu.get("WRH_GateInTime")),
            "RTU Unload Complete": parse_dt(rtu.get("WRH_UnloadCompleteTime")),
            "Receive Gate out Time": parse_dt(rtu.get("WRH_GateOutTime")),
            "Package ID": kph.get("KPH_PackageID"),
            "Quantity": p.get("KP_PackageQty"),
            "Package Type": pt_disp,
            "Weight": fmt_qty(p.get("KP_Weight"), p.get("KP_WeightUQ") or "KG"),
            "Length": fmt_qty(p.get("KP_Length"), p.get("KP_DimensionUQ") or "IN"),
            "Width": fmt_qty(p.get("KP_Width"), p.get("KP_DimensionUQ") or "IN"),
            "Height": fmt_qty(p.get("KP_Height"), p.get("KP_DimensionUQ") or "IN"),
            "Volume": fmt_qty(p.get("KP_Volume"), p.get("KP_VolumeUQ") or "M3"),
            "Location": loc.get("WLV_LocationString"),
            "Status": st_disp,
            "Commodity": p.get("KP_RH_NKCommodityCode") or None,
            "Description": p.get("KP_GoodsDescription") or None,
            "Marks and Numbers": p.get("KP_MarksAndNumbers") or None,
            # No consignment -> not allocated to a dispatch load list -> not ready to stage.
            "Is Ready to Stage": "N",
        })
        out.append(row)
    return out

def build_workbook(rows, path):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    wide = {"Receive Transportation Unit ID": 26, "Receive Gate in Time": 17, "RTU Unload Complete": 17,
            "Receive Gate out Time": 18, "Package ID": 22, "Package Type": 16, "Location": 16, "Status": 16,
            "Description": 30, "Marks and Numbers": 22, "Booking Party": 40, "Consignor": 40, "Consignee": 40,
            "Expected Arrival at Warehouse": 26, "Expected Dispatch from Warehouse": 28}
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Transit Package"
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c) for c in COLUMNS])
    for ci, col in enumerate(COLUMNS, 1):
        L = get_column_letter(ci); ws.column_dimensions[L].width = wide.get(col, 12)
        if col in DATE_COLS:
            for cell in ws[L][1:]:
                cell.number_format = "dd-mmm-yy hh:mm"
    ws.freeze_panes = "A2"
    ws.add_table(Table(displayName="TransitPackage",
                       ref=f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}",
                       tableStyleInfo=TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)))
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True); wb.save(path)

def upload(local_path):
    for k, v in {"AZURE_TENANT_ID": AZ_TENANT, "AZURE_CLIENT_ID": AZ_CLIENT, "AZURE_CLIENT_SECRET": AZ_SECRET,
                 "SHAREPOINT_HOSTNAME": SP_HOST, "SHAREPOINT_SITE_PATH": SP_SITE}.items():
        if not v:
            sys.exit(f"FATAL: UPLOAD=true but {k} is not set")
    graph = "https://graph.microsoft.com/v1.0"
    body = urllib.parse.urlencode({"client_id": AZ_CLIENT, "client_secret": AZ_SECRET,
                                   "scope": "https://graph.microsoft.com/.default",
                                   "grant_type": "client_credentials"}).encode()
    tok = json.load(_http(f"https://login.microsoftonline.com/{AZ_TENANT}/oauth2/v2.0/token", data=body,
                          method="POST", headers={"Content-Type": "application/x-www-form-urlencoded"}))["access_token"]
    H = {"Authorization": "Bearer " + tok}
    site = json.load(_http(f"{graph}/sites/{SP_HOST}:{SP_SITE}", headers=H))
    drive = json.load(_http(f"{graph}/sites/{site['id']}/drive", headers=H))
    dest = "/".join(p for p in (SP_FOLDER, os.path.basename(local_path)) if p)
    url = f"{graph}/drives/{drive['id']}/root:/{urllib.parse.quote(dest)}:/content"
    res = json.load(_http(url, data=open(local_path, "rb").read(), method="PUT",
                          headers={**H, "Content-Type": "application/octet-stream"}))
    log.info("Uploaded: %s (%s bytes)", res.get("webUrl"), res.get("size"))

def main():
    logging.basicConfig(level=logging.INFO, stream=sys.stdout, format="%(asctime)s %(levelname)s %(message)s")
    cutoff = days_ago_boundary(GATE_DAYS)
    log.info("PKG No-Consignment Report - branches=%s, gate-in since %s, no consignment", BRANCH_CODES, cutoff)
    cw = CargoWise(); cw.login()
    ptmap = cw.packtype_map()
    all_rows = []
    for code in BRANCH_CODES:
        recs = cw.pull_packages(cw.branch_map[code], cutoff)
        rows = shape(code, recs, ptmap)
        log.info("  %s: %d packages", code, len(rows))
        all_rows.extend(rows)
    out_path = os.path.join(OUTDIR, OUT_FILE)
    build_workbook(all_rows, out_path)
    log.info("Workbook written: %s (%d rows)", out_path, len(all_rows))
    if DO_UPLOAD:
        import sp_upload; sp_upload.upload(out_path, SP_FOLDER)
    else:
        log.info("UPLOAD disabled; skipping SharePoint upload")
    log.info("Done.")

if __name__ == "__main__":
    main()
