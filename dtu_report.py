#!/usr/bin/env python3
"""
AWA Data Services - DTU (Dispatch Transportation Units) Report - CCC
====================================================================
The dispatch-side mirror of rtu_report.py: pulls Dispatch Transportation Units
from CargoWise TWD for the CCC branch and writes the same column structure as the
RTU (KPIU) report, then (optionally) uploads to SharePoint via the shared uploader.

Filter (same shape as the RTU report): WDH_GateInTime on/after DTU_GATE_FROM and the
unit has at least one package with a status. CCC branch only.

Config is environment-driven (see .env.example). No secrets are stored in code.
"""
import os, sys, json, re, logging, datetime
import urllib.request, urllib.parse, urllib.error, http.cookiejar

def _load_dotenv(path=".env"):
    if os.path.exists(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())
_load_dotenv()

def _env(key, default=None, required=False):
    val = os.environ.get(key, default)
    if required and not val:
        sys.exit(f"FATAL: required environment variable {key} is not set")
    return val

CW_BASE      = _env("CW_BASE_URL", "https://www-isbint.wisegrid.net").rstrip("/")
CW_MODULE    = _env("CW_MODULE", "TWD")
ODATA_MODEL  = _env("CW_ODATA_MODEL", "TransitWarehouse")
CW_USERNAME  = _env("CW_USERNAME", required=True)
CW_PASSWORD  = _env("CW_PASSWORD", required=True)
BRANCH_CODES = [b.strip().upper() for b in _env("DTU_BRANCH_CODES", "CCC").split(",") if b.strip()]
DEPT_CODE    = _env("CW_DEPARTMENT_CODE", "BRN").upper()
GATE_FROM    = _env("DTU_GATE_FROM", "2024-09-04")
DISPLAY_TZ_OFFSET = float(_env("CC_DISPLAY_TZ_OFFSET", "-6"))

OUTDIR       = _env("OUTPUT_DIR", "./output")
OUT_FILE     = _env("DTU_FILENAME", "DTU.xlsx")
DO_UPLOAD    = _env("UPLOAD", "true").lower() in ("1", "true", "yes")
SP_FOLDER    = _env("DTU_SHAREPOINT_FOLDER", "DTU Report")

COLUMNS = ["Branch", "Dispatch Transportation Unit ID", "DTU Reference", "Transport Company",
           "Drivers Name", "Loaded Packages", "Load Complete", "Gate in Time"] + [f"Column{i}" for i in range(1, 9)]
DATE_COLS = {"Load Complete", "Gate in Time"}

log = logging.getLogger("dtu_report")
_cookies = http.cookiejar.CookieJar()
_opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_cookies))

def _http(url, data=None, headers=None, method=None, timeout=180):
    return _opener.open(urllib.request.Request(url, data=data, headers=headers or {}, method=method), timeout=timeout)

def parse_dt(s):
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})(?::(\d{2}))?(?:\.\d+)?(Z|[+-]\d{2}:?\d{2})?", s)
    if not m:
        return None
    y, mo, d, h, mi = (int(x) for x in m.groups()[:5])
    se = int(m.group(6) or 0); off = m.group(7)
    try:
        dt = datetime.datetime(y, mo, d, h, mi, se)
    except ValueError:
        return None
    if off and off != "Z":
        sign = 1 if off[0] == "+" else -1
        dt -= datetime.timedelta(hours=sign * int(off[1:3]), minutes=sign * int(off[-2:]))
    dt += datetime.timedelta(hours=DISPLAY_TZ_OFFSET)
    return dt.replace(second=0, microsecond=0)

class CargoWise:
    def __init__(self):
        self.auth = f"{CW_BASE}/Glow/auth/v2"; self.odata = f"{CW_BASE}/Glow/odata/{ODATA_MODEL}"
        self.user_key = None; self.branch_map = {}; self.dept_key = None

    def _post(self, path, body):
        r = _http(f"{self.auth}/{path}", data=json.dumps(body).encode(), method="POST",
                  headers={"Content-Type": "application/json", "Accept": "application/json",
                           "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop", "Origin": CW_BASE})
        return json.load(r)

    def _get(self, resource, params):
        url = f"{self.odata}/{resource}?{urllib.parse.urlencode(params)}"
        return json.load(_http(url, headers={"Accept": "application/json", "wtg-app": "Glow",
                                             "Referer": f"{CW_BASE}/{CW_MODULE}/Desktop"}))

    def login(self):
        res = self._post("credential/claim/Staff",
                         {"userName": CW_USERNAME, "password": CW_PASSWORD, "setTokenCookie": True})
        if res.get("result") != 0:
            sys.exit(f"FATAL: CargoWise login failed (result={res.get('result')})")
        self.user_key = res["userKey"]
        ctx = self._post("credential/context/list",
                         {"logonProviderType": "Staff", "userKey": self.user_key, "useTokenCookie": True})
        for b in ctx.get("branchInfos", []):
            if b["code"].upper() in BRANCH_CODES:
                self.branch_map[b["code"].upper()] = b["key"]
        for d in ctx.get("departmentInfos", []):
            if d["code"].upper() == DEPT_CODE:
                self.dept_key = d["key"]
        missing = [c for c in BRANCH_CODES if c not in self.branch_map]
        if missing:
            sys.exit(f"FATAL: branch code(s) not found: {missing}")
        if not self.dept_key:
            sys.exit(f"FATAL: department {DEPT_CODE} not found")
        self._select(self.branch_map[BRANCH_CODES[0]])
        log.info("Authenticated as %s", res.get("userDisplayName"))

    def _select(self, branch_key):
        self._post("credential/context/select",
                   {"logonProviderType": "Staff", "userKey": self.user_key,
                    "branchKey": branch_key, "departmentKey": self.dept_key, "useAndSetTokenCookie": True})
        self._post("session/begin", {"tokenType": 1, "sessionType": "General", "useAndSetTokenCookie": True})

    def country_map(self):
        m, skip = {}, 0
        try:
            while True:
                d = self._get("RefCountryInfos", {"$select": "RN_Code,RN_Desc", "$top": "50", "$skip": str(skip)})
                v = d.get("value", [])
                m.update({r["RN_Code"]: r["RN_Desc"] for r in v})
                if len(v) < 50:
                    break
                skip += 50
        except Exception as e:
            log.warning("Could not load country map (%s)", e)
        return m

    def pull_dtus(self, branch_key):
        self._select(branch_key)
        boundary_h = int(round(-DISPLAY_TZ_OFFSET))   # -6 -> 06:00Z
        filt = (f"WDH_GateInTime ge {GATE_FROM}T{boundary_h:02d}:00:00Z and "
                f"WhsItemPackageStates/any(p: p/WPS_Status ne '')")
        expand = ("Addresses($expand=Address($expand=OrgHeader)),"
                  "WhsItemPackageStates($count=true;$top=0)")
        sel = "WDH_ReferenceNumber,WDH_VehicleReference,WDH_SignedBy,WDH_GateInTime,WDH_LoadCompleteTime"
        rows, skip, page = [], 0, 50
        while True:
            params = {"$filter": filt, "$expand": expand, "$select": sel,
                      "$top": str(page), "$skip": str(skip), "$orderby": "WDH_ReferenceNumber"}
            for _ in range(4):
                try:
                    d = self._get("WhsItemDispatchTransportationUnits", params); break
                except urllib.error.HTTPError as e:
                    if e.code == 401:
                        log.info("  session expired, re-authenticating"); self.login(); self._select(branch_key); continue
                    raise
            else:
                raise RuntimeError("repeated failures pulling DTUs")
            batch = d.get("value", [])
            rows.extend(batch)
            if len(batch) < page:
                break
            skip += page
            if skip % 1000 == 0:
                log.info("  ...%d", len(rows))
        return rows

def transport_company(dtu, cmap):
    for a in dtu.get("Addresses", []):
        if a.get("E2_AddressType") == "TRA":
            ad = a.get("Address") or {}; oh = ad.get("OrgHeader") or {}
            name = a.get("E2_CompanyName") or oh.get("OH_FullName") or ad.get("OA_CompanyNameOverride") or ""
            a1 = a.get("E2_Address1") or ad.get("OA_Address1") or ad.get("OA_Code") or ""
            a2 = a.get("E2_Address2") or ad.get("OA_Address2") or ""
            city = a.get("E2_City") or ad.get("OA_City") or ""
            state = a.get("E2_State") or ad.get("OA_State") or ""
            pc = a.get("E2_Postcode") or ad.get("OA_PostCode") or ""
            cc = a.get("E2_RN_NKCountryCode") or ad.get("OA_RN_NKCountryCode") or ""
            country = cmap.get(cc, cc)
            csz = " ".join(x for x in (city, state, pc) if x)
            return ", ".join(x for x in (name, a1, a2, csz, country) if x)
    return ""

def shape(branch, records, cmap):
    out = []
    for r in records:
        out.append({
            "Branch": branch,
            "Dispatch Transportation Unit ID": r.get("WDH_ReferenceNumber"),
            "DTU Reference": r.get("WDH_VehicleReference"),
            "Transport Company": transport_company(r, cmap),
            "Drivers Name": r.get("WDH_SignedBy"),
            "Loaded Packages": r.get("WhsItemPackageStates@odata.count", 0),
            "Load Complete": parse_dt(r.get("WDH_LoadCompleteTime")),
            "Gate in Time": parse_dt(r.get("WDH_GateInTime")),
            **{f"Column{i}": None for i in range(1, 9)},
        })
    return out

def build_workbook(rows, path):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    widths = {"Branch": 8, "Dispatch Transportation Unit ID": 26, "DTU Reference": 26, "Transport Company": 55,
              "Drivers Name": 22, "Loaded Packages": 16, "Load Complete": 17, "Gate in Time": 17}
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Dispatch Transportation Unit"
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c) for c in COLUMNS])
    for ci, col in enumerate(COLUMNS, 1):
        L = get_column_letter(ci); ws.column_dimensions[L].width = widths.get(col, 11)
        if col in DATE_COLS:
            for cell in ws[L][1:]:
                cell.number_format = "dd-mmm-yy hh:mm"
    ws.freeze_panes = "A2"
    ws.add_table(Table(displayName="DispatchTU",
                       ref=f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}",
                       tableStyleInfo=TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)))
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True); wb.save(path)

def main():
    logging.basicConfig(level=logging.INFO, stream=sys.stdout, format="%(asctime)s %(levelname)s %(message)s")
    log.info("DTU Report - branches=%s, gate-in from %s, package status not blank", BRANCH_CODES, GATE_FROM)
    cw = CargoWise(); cw.login()
    cmap = cw.country_map()
    all_rows = []
    for code in BRANCH_CODES:
        recs = cw.pull_dtus(cw.branch_map[code])
        rows = shape(code, recs, cmap)
        log.info("  %s: %d DTUs", code, len(rows))
        all_rows.extend(rows)
    out_path = os.path.join(OUTDIR, OUT_FILE)
    build_workbook(all_rows, out_path)
    log.info("Workbook written: %s (%d rows)", out_path, len(all_rows))
    if DO_UPLOAD:
        import sp_upload; sp_upload.upload(out_path, SP_FOLDER)
    else:
        log.info("UPLOAD disabled; skipping SharePoint upload")
    log.info("Done.")

if __name__ == "__main__":
    main()
